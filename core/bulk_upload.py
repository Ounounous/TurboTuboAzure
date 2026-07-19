"""
Framework generico de carga masiva por Excel: valida TODO el archivo antes de guardar nada.
Si hay 1+ error, no se guarda nada y se descarga el mismo Excel con una columna ERRORES agregada
(fila por fila, resaltada) -- nunca queda una carga a medias. Reutilizado por leads, telefonos,
correos, aval, direccion, bienes, gestiones y las acciones de suspensiones.
"""
from io import BytesIO

import openpyxl
from django.http import HttpResponse
from openpyxl.styles import Font, PatternFill

from .text_utils import normalize_header

ERROR_FILL = PatternFill('solid', fgColor='F1D6D6')
HEADER_FILL = PatternFill('solid', fgColor='C0392B')
HEADER_FONT = Font(bold=True, color='FFFFFF')


class CargaMasivaResult:
    """`filas`: lista de dicts limpios, listos para guardar (dentro de un transaction.atomic()).
    `respuesta_error`: HttpResponse con el Excel de errores, listo para devolver tal cual --
    no se guardo nada. Exactamente uno de los dos esta seteado."""

    def __init__(self, filas=None, respuesta_error=None):
        self.filas = filas
        self.respuesta_error = respuesta_error

    @property
    def ok(self):
        return self.respuesta_error is None


def detectar_columnas(header_row, alias_map):
    """alias_map: {alias_normalizado: clave_interna}. Devuelve {clave_interna: indice_columna}
    (la primera columna que matchea cada clave, para no pisarse si el Excel repite un alias)."""
    colmap = {}
    for idx, raw in enumerate(header_row):
        key = alias_map.get(normalize_header(raw))
        if key and key not in colmap:
            colmap[key] = idx
    return colmap


def procesar_carga(excel_file, alias_map, columnas_requeridas, validar_fila, nombre_archivo='errores.xlsx'):
    """
    - excel_file: archivo subido (request.FILES).
    - alias_map: {alias_normalizado_de_columna: clave_interna} para detectar encabezados,
      tolerante a variantes de nombre/idioma/orden de columnas.
    - columnas_requeridas: claves internas que deben existir en el Excel.
    - validar_fila(fila: dict clave->valor crudo, numero_fila: int) -> (dict_limpio|None, [errores]).
      Debe validar TODO lo que se pueda saber sin tocar la base para escribir (existencia del
      lead, choices validas, formatos de fecha, etc.) y devolver los datos ya listos para usar.

    Devuelve CargaMasivaResult. Nunca escribe nada en la base -- eso es responsabilidad de quien
    llama, dentro de un transaction.atomic(), y solo si `resultado.ok`.
    """
    try:
        wb = openpyxl.load_workbook(excel_file, data_only=True)
    except Exception:
        return CargaMasivaResult(respuesta_error=_error_simple(
            'No se pudo leer el archivo. Debe ser un Excel (.xlsx) válido.'
        ))

    sheet = wb.active
    header_row = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True), ())
    colmap = detectar_columnas(header_row, alias_map)

    faltan = [c for c in columnas_requeridas if c not in colmap]
    if faltan:
        return CargaMasivaResult(respuesta_error=_error_simple(
            'Faltan columnas obligatorias en el Excel: ' + ', '.join(faltan) + '.'
        ))

    filas_limpias = []
    errores_por_fila = {}
    hay_datos = False

    for rownum, raw_row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
        if not raw_row or all(v in (None, '') for v in raw_row):
            continue
        hay_datos = True
        fila = {clave: (raw_row[idx] if idx < len(raw_row) else None) for clave, idx in colmap.items()}
        limpio, errores = validar_fila(fila, rownum)
        if errores:
            errores_por_fila[rownum] = errores
        else:
            filas_limpias.append(limpio)

    if not hay_datos:
        return CargaMasivaResult(respuesta_error=_error_simple('El archivo no tenía filas de datos.'))

    if errores_por_fila:
        return CargaMasivaResult(
            respuesta_error=_excel_de_errores(sheet, header_row, errores_por_fila, nombre_archivo)
        )

    return CargaMasivaResult(filas=filas_limpias)


def _excel_de_errores(sheet, header_row, errores_por_fila, nombre_archivo):
    """El mismo Excel de entrada + columna ERRORES al final de cada fila (opcion A)."""
    wb_out = openpyxl.Workbook()
    ws = wb_out.active
    ws.title = 'Errores'

    headers = [('' if h is None else str(h)) for h in header_row] + ['ERRORES']
    ws.append(headers)
    for cell in ws[1]:
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL

    for rownum, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
        if not row or all(v in (None, '') for v in row):
            continue
        vals = list(row)
        errores = errores_por_fila.get(rownum)
        vals.append('; '.join(errores) if errores else '')
        ws.append(vals)
        if errores:
            for cell in ws[ws.max_row]:
                cell.fill = ERROR_FILL

    for col in ws.columns:
        largo = max((len(str(c.value)) if c.value is not None else 0) for c in col)
        ws.column_dimensions[col[0].column_letter].width = min(max(largo + 2, 10), 60)

    return _responder_excel(wb_out, nombre_archivo)


def _error_simple(mensaje):
    """Para errores previos a poder leer filas (archivo invalido, columnas faltantes): se
    mantiene el mismo contrato (siempre se descarga un Excel, nunca solo un mensaje flash)."""
    wb_out = openpyxl.Workbook()
    ws = wb_out.active
    ws.title = 'Error'
    ws.append(['ERROR'])
    ws['A1'].font = HEADER_FONT
    ws['A1'].fill = HEADER_FILL
    ws.append([mensaje])
    ws['A2'].fill = ERROR_FILL
    ws.column_dimensions['A'].width = 100
    return _responder_excel(wb_out, 'errores.xlsx')


def _responder_excel(workbook, nombre_archivo):
    buf = BytesIO()
    workbook.save(buf)
    buf.seek(0)
    response = HttpResponse(
        buf.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = f'attachment; filename="{nombre_archivo}"'
    return response
