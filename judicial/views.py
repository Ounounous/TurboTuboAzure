from django.contrib import messages
from django.contrib.auth.mixins import LoginRequiredMixin
from django.core.exceptions import PermissionDenied
from django.shortcuts import get_object_or_404, redirect, render
from django.views import View
from openpyxl import load_workbook

from demographics.views import _normalize_header, find_lead
from lead.models import Lead

from .models import JudicialSettings

SUPERVISOR_TYPES = ('admin', 'owner', 'supervisor')


class SupervisorRequiredMixin(LoginRequiredMixin):
    def dispatch(self, request, *args, **kwargs):
        userprofile = getattr(request.user, 'userprofile', None)
        if not userprofile or userprofile.user_type not in SUPERVISOR_TYPES:
            raise PermissionDenied
        return super().dispatch(request, *args, **kwargs)


class JudicialHomeView(SupervisorRequiredMixin, View):
    template_name = 'judicial/index.html'

    def get(self, request, *args, **kwargs):
        context = {
            'config': JudicialSettings.get_solo(),
            'estado_judicial_choices': Lead.CHOICES_ESTADO_JUDICIAL,
        }
        return render(request, self.template_name, context)


class ToggleJudicialInfoView(SupervisorRequiredMixin, View):
    def post(self, request, *args, **kwargs):
        config = JudicialSettings.get_solo()
        config.mostrar_info_judicial = not config.mostrar_info_judicial
        config.updated_by = request.user
        config.save()
        if config.mostrar_info_judicial:
            messages.success(request, 'Información judicial: ahora se muestra en la ficha del cliente.')
        else:
            messages.success(request, 'Información judicial: ahora está oculta en la ficha del cliente.')
        return redirect('judicial:index')


class UpdateLeadEstadoJudicialView(SupervisorRequiredMixin, View):
    """Edición manual, un lead a la vez, desde su ficha."""

    def post(self, request, lead_id, *args, **kwargs):
        lead = get_object_or_404(Lead, pk=lead_id)
        valor = request.POST.get('estado_judicial', '')
        valid_values = dict(Lead.CHOICES_ESTADO_JUDICIAL)
        if valor in valid_values:
            lead.estado_judicial = valor
            lead.save(update_fields=['estado_judicial'])
            messages.success(request, f'Estado judicial de {lead.op} actualizado a "{valid_values[valor]}".')
        else:
            messages.error(request, 'Estado judicial inválido.')
        return redirect('leads:detail', pk=lead.pk)


class BulkUploadEstadoJudicialView(SupervisorRequiredMixin, View):
    """Carga masiva por Excel: columnas CARTERA, SUBCARTERA, ID, ESTADO_JUDICIAL."""

    def post(self, request, *args, **kwargs):
        excel_file = request.FILES.get('excel_file')
        if not excel_file:
            messages.error(request, 'Debes subir un archivo Excel.')
            return redirect('judicial:index')

        wb = load_workbook(excel_file)
        sheet = wb.active
        valid_values = dict(Lead.CHOICES_ESTADO_JUDICIAL)

        updated, errors = 0, []
        for row_number, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
            cartera, subcartera, op, estado = (list(row) + [None] * 4)[:4]
            if not cartera or not subcartera or not op or not estado:
                errors.append(f"Fila {row_number}: faltan datos (cartera, subcartera, ID o estado judicial).")
                continue

            lead = find_lead(cartera, subcartera, op)
            if not lead:
                errors.append(f"Fila {row_number}: no se encontró lead OP={op} en Cartera={cartera}, Subcartera={subcartera}.")
                continue

            estado_normalizado = _normalize_header(estado)
            if estado_normalizado not in valid_values:
                errors.append(f"Fila {row_number}: estado judicial '{estado}' no es válido.")
                continue

            lead.estado_judicial = estado_normalizado
            lead.save(update_fields=['estado_judicial'])
            updated += 1

        if updated:
            messages.success(request, f'Se actualizó el estado judicial de {updated} lead(s).')
        for error in errors[:20]:
            messages.error(request, error)
        if len(errors) > 20:
            messages.error(request, f'... y {len(errors) - 20} error(es) más.')

        return redirect('judicial:index')
