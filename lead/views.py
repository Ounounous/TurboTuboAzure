from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.contrib.auth.mixins import LoginRequiredMixin
from django.db import transaction
from django.db.models import Count, F, Max, Q
from django.core.exceptions import PermissionDenied
from django.shortcuts import redirect, get_object_or_404, render
from django.urls import reverse, reverse_lazy
from django.http import HttpResponse, JsonResponse
from django.views.generic import ListView, DetailView, DeleteView, UpdateView, CreateView, View
from django.utils.timezone import now, localtime, localdate
from .models import StatusChangeLog, Team
import openpyxl
from io import BytesIO
import logging

logger = logging.getLogger(__name__)

from openpyxl import load_workbook

from .forms import AddFileForm, AddLeadForm, UploadExcelFileForm, UploadAssignmentFileForm, QuickAssignForm
from .models import Lead, LeadAssignment, LeadFile, LeadNote, User
from core.concurrency import con_limite_concurrencia
from cartera.models import Cartera, Subcartera
from demographics.models import IDDemographics, AvalDemographics, IDItem, Phone


class _SupervisorGate(LoginRequiredMixin):
    """Solo admin/owner/supervisor. Subir/descargar cartera no es para cobradores."""
    def dispatch(self, request, *args, **kwargs):
        from .permissions import es_supervisor
        if not es_supervisor(request.user):
            raise PermissionDenied
        return super().dispatch(request, *args, **kwargs)


class _AdminOwnerGate(LoginRequiredMixin):
    """Solo admin/owner (ej. eliminar clientes)."""
    def dispatch(self, request, *args, **kwargs):
        from .permissions import es_admin_owner
        if not es_admin_owner(request.user):
            raise PermissionDenied
        return super().dispatch(request, *args, **kwargs)


class LeadListView(LoginRequiredMixin, ListView):
    model = Lead

    # Filtros por columna (además del buscador general por ID/RUT). Nombre del parámetro GET ->
    # lookup en el queryset. Pensado para ser expandible: agregar una columna es una línea acá.
    COLUMN_FILTERS = {
        'op': 'op__icontains',
        'name': 'name__icontains',
        'rut': 'rut__icontains',
        'status': 'status',
        'ciclo': 'ciclo',
        'ciclo_cartera': 'ciclo_cartera',
        'activo': 'activo',
        'tipo_cobranza': 'tipo_cobranza',
        'tiene_aval': 'tiene_aval',
        'cartera': 'subcartera__cartera__nombre__icontains',
        'subcartera': 'subcartera__nombre__icontains',
    }

    # Columnas ordenables: clave del parámetro ?sort= -> campo del modelo.
    SORT_FIELDS = {
        'op': 'op', 'nombre': 'name', 'rut': 'rut',
        'insoluto': 'saldo_insoluto', 'deuda': 'saldo_deuda', 'cuota': 'valor_cuota',
        'atrasadas': 'cuotas_atrasadas', 'status': 'status',
        'cartera': 'subcartera__cartera__nombre', 'dias': 'last_action_at',
    }

    def get_limit(self):
        try:
            limit = int(self.request.GET.get('limit', 10))
        except ValueError:
            limit = 10
        return limit if limit in (10, 50, 100) else 10

    def _base_scope(self):
        """Leads visibles para el usuario segun su rol (ver lead/permissions.py):
        cobrador = sus asignados/creados; supervisor = sus carteras; admin/owner = todo."""
        from .permissions import leads_visibles
        return leads_visibles(self.request.user)

    def get_queryset(self):
        queryset = self._base_scope().select_related('subcartera__cartera').annotate(
            # Última gestión (Action) del lead — NO cuenta las notas (LeadNote es otro modelo).
            last_action_at=Max('actions__created_at')
        )

        # Solo favoritos del usuario.
        if self.request.GET.get('fav') == '1':
            queryset = queryset.filter(favorited_by=self.request.user)

        # Buscador general: ID (op), nombre y RUT en un solo campo.
        q = self.request.GET.get('q', '').strip()
        if q:
            queryset = queryset.filter(
                Q(op__icontains=q) | Q(name__icontains=q) | Q(rut__icontains=q)
            )

        # Filtros por columna (panel avanzado, expandible).
        for param, lookup in self.COLUMN_FILTERS.items():
            value = self.request.GET.get(param, '').strip()
            if value:
                queryset = queryset.filter(**{lookup: value})

        return self._apply_sort(queryset)

    def _apply_sort(self, queryset):
        sort = self.request.GET.get('sort', 'nombre')
        direction = self.request.GET.get('dir', 'asc')
        if sort not in self.SORT_FIELDS:
            sort = 'nombre'
        if sort == 'dias':
            # "días desde gestión" = inverso de last_action_at (más antiguo = más días).
            # asc (menos días primero) => last_action_at desc; nunca gestionado (null) al final.
            campo = F('last_action_at')
            order = campo.desc(nulls_last=True) if direction == 'asc' else campo.asc(nulls_last=True)
            return queryset.order_by(order, 'name')
        field = self.SORT_FIELDS[sort]
        return queryset.order_by(field if direction == 'asc' else '-' + field, 'name')

    def _url_with(self, **changes):
        """URL de la misma vista cambiando algunos parámetros (None = quitar)."""
        params = self.request.GET.copy()
        for key, value in changes.items():
            if value is None:
                params.pop(key, None)
            else:
                params[key] = value
        encoded = params.urlencode()
        return ('?' + encoded) if encoded else '?'

    def _sort_links(self):
        cur_sort = self.request.GET.get('sort', 'nombre')
        cur_dir = self.request.GET.get('dir', 'asc')
        links = {}
        for key in self.SORT_FIELDS:
            new_dir = 'desc' if (cur_sort == key and cur_dir == 'asc') else 'asc'
            arrow = ('▲' if cur_dir == 'asc' else '▼') if cur_sort == key else ''
            links[key] = {'url': self._url_with(sort=key, dir=new_dir), 'arrow': arrow}
        return links

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        limit = self.get_limit()
        base_qs = self.get_queryset()
        context['total_count'] = base_qs.count()

        # ids favoritos del usuario (para pintar la estrella) sobre la página mostrada.
        leads = list(base_qs[:limit])
        fav_ids = set(
            self._base_scope().filter(favorited_by=self.request.user, pk__in=[l.pk for l in leads])
            .values_list('pk', flat=True)
        )
        today = localdate()
        for lead in leads:
            lead.is_fav = lead.pk in fav_ids
            if lead.last_action_at:
                lead.dias_ultima_gestion = (today - localtime(lead.last_action_at).date()).days
            else:
                lead.dias_ultima_gestion = None
        context['leads'] = leads
        context['limit'] = limit

        # Conteos por status (chips) sobre el scope del usuario, sin aplicar el filtro de status.
        scope = self._base_scope()
        status_counts = {row['status']: row['n'] for row in scope.values('status').annotate(n=Count('id'))}
        selected_status = self.request.GET.get('status', '')
        context['status_chips'] = [
            {'value': val, 'label': label, 'count': status_counts.get(val, 0),
             'color': Lead.STATUS_COLOR.get(val, 'slate'),
             'url': self._url_with(status=val), 'active': selected_status == val}
            for val, label in Lead.CHOICES_STATUS
        ]
        context['total_scope'] = scope.count()
        context['fav_count'] = scope.filter(favorited_by=self.request.user).count()
        context['url_todos'] = self._url_with(status=None, fav=None)
        context['url_fav'] = self._url_with(fav='1', status=None)
        context['url_clear_fav'] = self._url_with(fav=None)
        context['sort_links'] = self._sort_links()
        context['url_limit_10'] = self._url_with(limit=10)
        context['url_limit_50'] = self._url_with(limit=50)
        context['url_limit_100'] = self._url_with(limit=100)

        # Estado de filtros/orden para la UI.
        context['q'] = self.request.GET.get('q', '')
        context['selected_status'] = selected_status
        context['only_fav'] = self.request.GET.get('fav') == '1'
        context['sort'] = self.request.GET.get('sort', 'nombre')
        context['dir'] = self.request.GET.get('dir', 'asc')
        active_filters = {p: self.request.GET.get(p, '') for p in self.COLUMN_FILTERS}
        context['filters'] = active_filters
        context['advanced_open'] = any(v for k, v in active_filters.items() if k != 'status')

        context['status_choices'] = Lead.CHOICES_STATUS
        context['ciclo_choices'] = Lead.CHOICES_CICLO
        context['ciclo_cartera_choices'] = Lead.CHOICES_CICLO_CARTERA
        context['activo_choices'] = Lead.CHOICES_ACTIVO
        context['tipo_cobranza_choices'] = Lead.CHOICES_TIPO_COBRANZA
        context['aval_choices'] = Lead.CHOICES_AVAL
        from .permissions import es_supervisor
        context['is_supervisor'] = es_supervisor(self.request.user)
        return context

class LeadDetailView(LoginRequiredMixin, DetailView):
    model = Lead

    def get_queryset(self):
        # Alcance por rol (cobrador: suyos; supervisor: sus carteras; admin/owner: todo).
        from .permissions import leads_visibles
        return leads_visibles(self.request.user, base=super().get_queryset())

    def get(self, request, *args, **kwargs):
        response = super().get(request, *args, **kwargs)
        # Registro de acceso a datos del deudor (Ley 20.575): quien vio la ficha y cuando.
        from configuracion.models import AccessLog, registrar_acceso
        registrar_acceso(request.user, AccessLog.VER_FICHA, lead=self.object)
        return response

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['demographics'] = IDDemographics.objects.filter(lead=self.object)
        context['aval_demographics'] = AvalDemographics.objects.filter(
            id_demographics__in=context['demographics']).first()
        context['fileform'] = AddFileForm()
        context['notes'] = self.object.notes.select_related('author')
        context['phones'] = self.object.phone_set.all()

        # Días desde la última gestión (no cuenta notas), igual que en la lista de clientes.
        last_action = self.object.actions.order_by('-created_at').first()
        if last_action:
            context['dias_ultima_gestion'] = (localdate() - localtime(last_action.created_at).date()).days
        else:
            context['dias_ultima_gestion'] = None

        # Próximo compromiso de pago (el más cercano hacia adelante; si no hay futuro, el último).
        today = localdate()
        next_commitment = self.object.payment_commitments.filter(fecha_compromiso__gte=today).order_by('fecha_compromiso').first()
        context['next_commitment'] = next_commitment or self.object.payment_commitments.order_by('-fecha_compromiso').first()

        # Iniciales para el avatar de la ficha.
        parts = self.object.name.split()
        context['initials'] = ((parts[0][0] if parts else '?') + (parts[1][0] if len(parts) > 1 else '')).upper()

        # Acciones de supervisor (marcar al día, suspender/desasignar/reactivar).
        user_type = getattr(self.request.user.userprofile, 'user_type', '')
        context['is_supervisor'] = user_type in ('admin', 'owner', 'supervisor')

        # Formulario de asignación rápida (solo supervisores)
        if context['is_supervisor']:
            team = getattr(self.request.user.userprofile, 'active_team', None)
            context['quick_assign_form'] = QuickAssignForm(team=team)

        return context

    def post(self, request, *args, **kwargs):
        from .permissions import leads_visibles, es_supervisor
        if not es_supervisor(request.user):
            raise PermissionDenied

        lead = self.get_object()
        if not leads_visibles(request.user, base=Lead.objects.filter(pk=lead.pk)).exists():
            raise PermissionDenied

        team = getattr(request.user.userprofile, 'active_team', None)
        form = QuickAssignForm(request.POST, team=team)

        if form.is_valid():
            collector = form.cleaned_data['collector']
            lead.assigned_to = collector
            # Reasignar saca al lead de desasignado, lo vuelve activo
            if lead.activo == Lead.DESASIGNADO:
                lead.activo = Lead.ACTIVO
                lead.desasignado_at = None
            lead.save()
            LeadAssignment.objects.create(lead=lead, user=collector, assigned_by=request.user)
            messages.success(request, f'Cliente asignado a {collector.username}')
            return redirect('leads:detail', pk=lead.pk)

        return self.get(request, *args, **kwargs)
class LeadDeleteView(_AdminOwnerGate, DeleteView):
    model = Lead
    success_url = reverse_lazy('leads:list')

    def get(self, request, *args, **kwargs):
        return self.post(request, *args, **kwargs)

class MarcarAlDiaView(LoginRequiredMixin, View):
    """
    Unica forma manual de tocar el status de un lead: un supervisor confirma que la cuenta
    esta al dia. El resto del status se calcula solo a partir de gestiones y pagos
    (ver actions/status_logic.py) -- no es editable por colectores.
    """
    def post(self, request, pk, *args, **kwargs):
        user_type = getattr(request.user.userprofile, 'user_type', '')
        if user_type not in ('admin', 'owner', 'supervisor'):
            raise PermissionDenied

        # leads_visibles acota al supervisor a sus carteras (admin/owner sin restriccion) --
        # antes cualquier supervisor podia marcar "al dia" un lead de una cartera ajena por pk.
        from .permissions import leads_visibles
        lead = get_object_or_404(leads_visibles(request.user), pk=pk)
        from actions.status_logic import apply_status
        apply_status(lead, Lead.AL_DIA, changed_by=request.user)
        messages.success(request, f'{lead.op} marcado como al día.')
        return redirect('leads:detail', pk=lead.pk)

@login_required
def status_changes_by_date(request, period='day'):
    # Fecha de inicio en la zona horaria local (evita perder cambios de las últimas horas
    # por el desfase UTC: usamos localdate + el lookup __date que compara por día local).
    from .permissions import scope_por_lead

    today = localdate()
    start_date = today.replace(day=1) if period == 'month' else today

    logs = StatusChangeLog.objects.filter(
        timestamp__date__gte=start_date
    ).select_related('lead', 'changed_by').order_by('-timestamp')

    user_type = getattr(request.user.userprofile, 'user_type', '')
    if user_type == 'supervisor':
        # Supervisor: solo su(s) cartera(s) -- antes veia cambios de TODAS las carteras.
        logs = scope_por_lead(logs, request.user)
    elif user_type not in ('admin', 'owner'):
        # Cobrador: solo lo que el mismo cambio (ya acotado, sin tocar mas).
        logs = logs.filter(changed_by=request.user)

    return render(request, 'lead/status_changes_list.html', {'logs': logs, 'period': period})

class LeadCreateView(_SupervisorGate, CreateView):
    model = Lead
    form_class = AddLeadForm
    success_url = reverse_lazy('leads:list')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        team = self.request.user.userprofile.active_team
        context['team'] = team
        context['title'] = 'Agregar cliente'

        return context

    def form_valid(self, form):
        team = self.request.user.userprofile.active_team

        self.object = form.save(commit=False)
        self.object.created_by = self.request.user
        self.object.assigned_to = self.request.user
        self.object.team = self.request.user.userprofile.active_team
        self.object.save()

        return redirect(self.get_success_url())



# Encabezados aceptados para la carga de clientes: el MISMO formato que produce "Descargar
# Cartera" (bajar, corregir y resubir funciona directo). Las columnas se detectan por nombre,
# no por posicion, asi que el orden y las columnas extra no importan.
LEAD_UPLOAD_ALIASES = {
    'op': 'op', 'operacion': 'op', 'id': 'op',
    'name': 'name', 'nombre': 'name', 'cliente': 'name',
    'rut': 'rut',
    'dv': 'dv',
    'saldo_insoluto': 'saldo_insoluto', 'insoluto': 'saldo_insoluto',
    'saldo_deuda': 'saldo_deuda', 'deuda': 'saldo_deuda',
    'valor_cuota': 'valor_cuota', 'cuota': 'valor_cuota',
    'cuotas_atrasadas': 'cuotas_atrasadas', 'atrasadas': 'cuotas_atrasadas',
    'cartera': 'cartera',
    'subcartera': 'subcartera',
    'tipo_cobranza': 'tipo_cobranza',
    'ciclo_cartera': 'ciclo_cartera',
    'ciclo': 'ciclo',
    'activo': 'activo',
    'tiene_aval': 'tiene_aval', 'aval': 'tiene_aval',
    # 'status' se ignora a proposito: el status se calcula solo (actions/status_logic.py).
}

LEAD_UPLOAD_REQUERIDAS = (
    'op', 'name', 'rut', 'dv', 'saldo_insoluto', 'saldo_deuda', 'valor_cuota', 'cuotas_atrasadas',
)


def _parse_entero(valor, campo, errores, minimo=0):
    """Acepta numeros de Excel o texto con $ / separadores de miles. None si es invalido."""
    if valor in (None, ''):
        errores.append(f'{campo}: vacío')
        return None
    if isinstance(valor, (int, float)):
        return int(valor)
    texto = str(valor).strip().replace('$', '').replace('.', '').replace(',', '').replace(' ', '')
    try:
        numero = int(texto)
    except ValueError:
        errores.append(f'{campo}: "{valor}" no es un número')
        return None
    if numero < minimo:
        errores.append(f'{campo}: no puede ser negativo')
        return None
    return numero


def _parse_choice(valor, choices, default, campo, errores):
    """Vacío -> default. Acepta el valor interno o la etiqueta, sin importar mayúsculas/espacios."""
    if valor in (None, ''):
        return default
    v = str(valor).strip().lower().replace(' ', '')
    for interno, etiqueta in choices:
        if v in (interno.lower().replace(' ', ''), str(etiqueta).strip().lower().replace(' ', '')):
            return interno
    errores.append(f'{campo}: "{valor}" no es un valor válido ({", ".join(i for i, _ in choices)})')
    return default


class ClientesTemplateView(_SupervisorGate, View):
    """Plantilla Excel para la carga masiva de clientes (leads) de una cartera."""
    def get(self, request, *args, **kwargs):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Clientes'
        ws.append([
            'op', 'name', 'rut', 'dv', 'saldo_insoluto', 'saldo_deuda', 'valor_cuota',
            'cuotas_atrasadas', 'cartera', 'subcartera', 'tipo_cobranza', 'ciclo_cartera',
            'ciclo', 'activo', 'tiene_aval',
        ])
        ws.append([
            'OP-0001', 'Juan Pérez', '12345678', '9', 500000, 550000, 45000,
            2, 'CARTERA-EJEMPLO', 'SUBCARTERA-EJEMPLO', 'extra judicial', '2026-1',
            '1', 'activo', 'no',
        ])
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        response = HttpResponse(
            content=output.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        response['Content-Disposition'] = 'attachment; filename=plantilla_clientes.xlsx'
        return response


class UploadExcelFileView(_SupervisorGate, View):
    def get(self, request, *args, **kwargs):
        form = UploadExcelFileForm()
        return render(request, 'lead/upload_excel.html', {'form': form})

    def post(self, request, *args, **kwargs):
        from core.bulk_upload import procesar_carga

        form = UploadExcelFileForm(request.POST, request.FILES)
        if not form.is_valid():
            messages.error(request, 'Selecciona la cartera y el archivo Excel.')
            return redirect('leads:upload-excel')

        cartera = form.cleaned_data['cartera']
        excel_file = form.cleaned_data['excel_file']

        # Para validar duplicados sin una query por fila.
        ops_existentes = {
            op.strip().upper()
            for op in Lead.objects.filter(subcartera__cartera=cartera).values_list('op', flat=True)
        }
        ops_en_archivo = set()

        def validar_fila(fila, rownum):
            errores = []
            op = str(fila.get('op') or '').strip()
            if not op:
                errores.append('op: vacío')
            else:
                if op.upper() in ops_existentes:
                    errores.append(f'op: "{op}" ya existe en la cartera {cartera.nombre}')
                if op.upper() in ops_en_archivo:
                    errores.append(f'op: "{op}" está repetido en este archivo')
                ops_en_archivo.add(op.upper())

            name = str(fila.get('name') or '').strip()
            if not name:
                errores.append('nombre: vacío')

            rut = _parse_entero(fila.get('rut'), 'rut', errores, minimo=1)
            # OJO: no usar "fila.get('dv') or ''" -- 0 es un DV valido pero es falsy en Python,
            # asi que "0 or ''" da '' y un DV=0 legitimo se reportaba como invalido.
            dv_valor = fila.get('dv')
            dv = '' if dv_valor is None else str(dv_valor).strip().upper()
            if not dv or len(dv) > 1:
                errores.append(f'dv: "{dv_valor}" inválido (un dígito o K)')

            saldo_insoluto = _parse_entero(fila.get('saldo_insoluto'), 'saldo_insoluto', errores)
            saldo_deuda = _parse_entero(fila.get('saldo_deuda'), 'saldo_deuda', errores)
            valor_cuota = _parse_entero(fila.get('valor_cuota'), 'valor_cuota', errores)
            cuotas = _parse_entero(fila.get('cuotas_atrasadas'), 'cuotas_atrasadas', errores)

            # Si el Excel trae columna cartera, debe coincidir con la cartera elegida en el
            # formulario (evita subir la cartera equivocada de un archivo descargado).
            cartera_archivo = str(fila.get('cartera') or '').strip()
            if cartera_archivo and cartera_archivo.lower() != cartera.nombre.lower():
                errores.append(f'cartera: el archivo dice "{cartera_archivo}" pero elegiste "{cartera.nombre}"')

            tipo_cobranza = _parse_choice(
                fila.get('tipo_cobranza'), Lead.CHOICES_TIPO_COBRANZA, Lead.EXTRAJUDICIAL, 'tipo_cobranza', errores)
            ciclo_cartera = _parse_choice(
                fila.get('ciclo_cartera'), Lead.CHOICES_CICLO_CARTERA, Lead.VIGENTE, 'ciclo_cartera', errores)
            ciclo = _parse_choice(fila.get('ciclo'), Lead.CHOICES_CICLO, Lead.NO_DEFINIDO, 'ciclo', errores)
            activo = _parse_choice(fila.get('activo'), Lead.CHOICES_ACTIVO, Lead.ACTIVO, 'activo', errores)
            tiene_aval = _parse_choice(fila.get('tiene_aval'), Lead.CHOICES_AVAL, Lead.NO, 'tiene_aval', errores)

            if errores:
                return None, errores
            return {
                'op': op, 'name': name, 'rut': rut, 'dv': dv,
                'saldo_insoluto': saldo_insoluto, 'saldo_deuda': saldo_deuda,
                'valor_cuota': valor_cuota, 'cuotas_atrasadas': cuotas,
                'subcartera_nombre': str(fila.get('subcartera') or '').strip(),
                'tipo_cobranza': tipo_cobranza, 'ciclo_cartera': ciclo_cartera,
                'ciclo': ciclo, 'activo': activo, 'tiene_aval': tiene_aval,
            }, []

        resultado = procesar_carga(
            excel_file, LEAD_UPLOAD_ALIASES, LEAD_UPLOAD_REQUERIDAS, validar_fila,
            nombre_archivo=f'errores_clientes_{cartera.slug}.xlsx',
        )
        if not resultado.ok:
            # Nada se guardó: se devuelve el mismo Excel con la columna ERRORES.
            return resultado.respuesta_error

        # Todo válido: se guarda completo en una sola transacción.
        with transaction.atomic():
            sub_default = cartera.subcartera_default
            sub_cache = {}

            def subcartera_para(nombre):
                if not nombre:
                    return sub_default
                clave = nombre.lower()
                if clave not in sub_cache:
                    sub = Subcartera.objects.filter(cartera=cartera, nombre__iexact=nombre).first()
                    if not sub:
                        sub = Subcartera.objects.create(cartera=cartera, nombre=nombre)
                    sub_cache[clave] = sub
                return sub_cache[clave]

            team = request.user.userprofile.active_team
            nuevos = [
                Lead(
                    op=f['op'], name=f['name'], rut=f['rut'], dv=f['dv'],
                    saldo_insoluto=f['saldo_insoluto'], saldo_deuda=f['saldo_deuda'],
                    valor_cuota=f['valor_cuota'], cuotas_atrasadas=f['cuotas_atrasadas'],
                    subcartera=subcartera_para(f['subcartera_nombre']),
                    tipo_cobranza=f['tipo_cobranza'], ciclo_cartera=f['ciclo_cartera'],
                    ciclo=f['ciclo'], activo=f['activo'], tiene_aval=f['tiene_aval'],
                    created_by=request.user, assigned_to=request.user, team=team,
                )
                for f in resultado.filas
            ]
            Lead.objects.bulk_create(nuevos, batch_size=500)

        messages.success(request, f'{len(resultado.filas)} cliente(s) cargado(s) en {cartera.nombre}.')
        return redirect('leads:list')

class DownloadExcelView(_SupervisorGate, View):
    @con_limite_concurrencia('export_excel', slots=2)
    def get(self, request, *args, **kwargs):
        from .permissions import leads_visibles
        # Create your Excel file here
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = 'Clients'

        # Add headers
        headers = ['Op','Name', 'RUT', 'DV', 'Saldo Insoluto', 'Saldo Deuda', 'Valor Cuota', 'Cuotas Atrasadas', 'Cartera', 'Subcartera', 'Tipo Cobranza', 'Status', 'Ciclo Cartera', 'Ciclo', 'Activo', 'Tiene Aval']
        sheet.append(headers)

        # Add data (solo las carteras del usuario; admin/owner: todo)
        for lead in leads_visibles(request.user, base=Lead.objects.select_related('subcartera__cartera')):
            sheet.append([
                lead.op, lead.name, lead.rut, lead.dv, lead.saldo_insoluto, lead.saldo_deuda, lead.valor_cuota, lead.cuotas_atrasadas, lead.subcartera.cartera.nombre, lead.subcartera.nombre, lead.tipo_cobranza, lead.get_status_display(), lead.ciclo_cartera, lead.ciclo, lead.activo, lead.tiene_aval
            ])

        # Save the workbook to a BytesIO stream
        output = BytesIO()
        workbook.save(output)
        output.seek(0)

        # Create the HTTP response
        response = HttpResponse(
            content=output.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename=clients.xlsx'
        from configuracion.models import AccessLog, registrar_acceso
        registrar_acceso(request.user, AccessLog.EXPORTAR_CLIENTES, detail='Excel de clientes')
        return response

class AddFileView(LoginRequiredMixin, View):
    def post(self, request, *args, **kwargs):
        from .permissions import leads_visibles
        # Alcance: solo se puede adjuntar a un lead dentro del alcance del usuario (antes se
        # aceptaba cualquier pk).
        lead = get_object_or_404(leads_visibles(request.user), pk=kwargs.get('pk'))

        # Tope de archivos por lead.
        if LeadFile.objects.filter(lead=lead).count() >= LeadFile.MAX_POR_LEAD:
            messages.error(request, f'Máximo {LeadFile.MAX_POR_LEAD} archivos por cliente.')
            return redirect('leads:detail', pk=lead.pk)

        form = AddFileForm(request.POST, request.FILES)
        if form.is_valid():
            file = form.save(commit=False)
            file.team = request.user.userprofile.active_team
            file.lead = lead
            file.created_by = request.user
            file.save()
            messages.success(request, 'Archivo adjuntado.')
        else:
            for errores in form.errors.values():
                for error in errores:
                    messages.error(request, error)
        return redirect('leads:detail', pk=lead.pk)

def _puede_ver_lead(user, lead):
    """Un usuario puede ver/anotar un lead si esta dentro de su alcance (mismo criterio que el
    resto de la app): cobrador -> suyos; supervisor -> sus carteras (antes CUALQUIER supervisor
    pasaba, incluso de una cartera ajena); admin/owner -> todo."""
    from .permissions import leads_visibles
    return leads_visibles(user).filter(pk=lead.pk).exists()


class ToggleFavoriteView(LoginRequiredMixin, View):
    """Marca/desmarca un lead como favorito del usuario (por AJAX, sin recargar)."""

    def post(self, request, *args, **kwargs):
        lead = get_object_or_404(Lead, pk=kwargs.get('pk'))
        if not _puede_ver_lead(request.user, lead):
            raise PermissionDenied
        if lead.favorited_by.filter(pk=request.user.pk).exists():
            lead.favorited_by.remove(request.user)
            favorited = False
        else:
            lead.favorited_by.add(request.user)
            favorited = True
        if request.headers.get('x-requested-with') == 'XMLHttpRequest':
            return JsonResponse({'ok': True, 'favorited': favorited})
        return redirect(request.POST.get('next') or reverse('leads:list'))


class AddLeadNoteView(LoginRequiredMixin, View):
    """Crea una nota interna en un lead. Las notas NO entran en los reportes de cartera."""

    def post(self, request, *args, **kwargs):
        lead = get_object_or_404(Lead, pk=kwargs.get('pk'))
        if not _puede_ver_lead(request.user, lead):
            raise PermissionDenied

        is_ajax = request.headers.get('x-requested-with') == 'XMLHttpRequest'
        body = (request.POST.get('body') or '').strip()

        if not body:
            if is_ajax:
                return JsonResponse({'ok': False, 'error': 'La nota no puede estar vacía.'}, status=400)
            messages.error(request, 'La nota no puede estar vacía.')
            return redirect(request.POST.get('next') or reverse('leads:detail', kwargs={'pk': lead.pk}))

        note = LeadNote.objects.create(lead=lead, author=request.user, body=body)

        if is_ajax:
            return JsonResponse({
                'ok': True,
                'note': {
                    'author': request.user.username,
                    'created_at': localtime(note.created_at).strftime('%d-%m-%Y %H:%M'),
                    'body': note.body,
                },
            })

        messages.success(request, 'Nota agregada.')
        return redirect(request.POST.get('next') or reverse('leads:detail', kwargs={'pk': lead.pk}))


ASSIGNMENT_UPLOAD_ALIASES = {
    'cartera': 'cartera',
    'subcartera': 'subcartera',
    'op': 'op', 'operacion': 'op', 'id': 'op',
    'collector': 'collector', 'cobrador': 'collector', 'username': 'collector', 'usuario': 'collector',
}


class DownloadAssignmentTemplateView(LoginRequiredMixin, View):
    """Plantilla Excel para la asignacion masiva de leads (carga por cartera/subcartera/op/collector)."""

    def get(self, request, *args, **kwargs):
        from .permissions import es_supervisor
        if not es_supervisor(request.user):
            raise PermissionDenied

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Asignaciones'
        ws.append(['Cartera', 'Subcartera', 'OP', 'Collector'])
        ws.append(['CARTERA-EJEMPLO', 'SUBCARTERA-EJEMPLO', 'OP-EJEMPLO', 'nombre_usuario_cobrador'])

        output = BytesIO()
        wb.save(output)
        output.seek(0)
        response = HttpResponse(
            content=output.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        response['Content-Disposition'] = 'attachment; filename=plantilla_asignacion.xlsx'
        return response


class AssignLeadsView(LoginRequiredMixin, View):
    template_name = "lead/leads_assign.html"

    def get(self, request, *args, **kwargs):
        from .permissions import es_supervisor
        if not es_supervisor(request.user):
            raise PermissionDenied

        # La asignacion una-a-una (checkboxes) se quito a proposito: con 10.000 leads esa lista
        # se vuelve inusable, y la asignacion individual ya vive en el detalle del cliente
        # (reasignar a cualquier cobrador del equipo). Aca queda solo la asignacion masiva por Excel.
        upload_form = UploadAssignmentFileForm()
        return render(request, self.template_name, {'upload_form': upload_form})

    def post(self, request, *args, **kwargs):
        from .permissions import es_supervisor, leads_visibles
        if not es_supervisor(request.user):
            raise PermissionDenied

        team = request.user.userprofile.active_team
        upload_form = UploadAssignmentFileForm(request.POST, request.FILES)

        if upload_form.is_valid():
            from core.bulk_upload import procesar_carga

            def validar_fila(fila, rownum):
                errores = []
                cartera_nombre = str(fila.get('cartera') or '').strip()
                subcartera_nombre = str(fila.get('subcartera') or '').strip()
                op = str(fila.get('op') or '').strip()
                collector_username = str(fila.get('collector') or '').strip()

                cartera = Cartera.objects.filter(nombre__iexact=cartera_nombre).first() if cartera_nombre else None
                if not cartera:
                    errores.append(f'cartera "{cartera_nombre}" no encontrada')

                lead = None
                if cartera and op:
                    lead = Lead.objects.filter(op__iexact=op, subcartera__cartera=cartera, team=team).first()
                    if not lead:
                        errores.append(f'no se encontró el lead OP={op} en cartera {cartera_nombre}')
                    elif not leads_visibles(request.user, base=Lead.objects.filter(pk=lead.pk)).exists():
                        errores.append(f'no tienes permiso sobre la cartera {cartera_nombre} (OP={op})')

                collector_obj = None
                if collector_username:
                    collector_obj = User.objects.filter(
                        username__iexact=collector_username, userprofile__active_team=team
                    ).first()
                if not collector_obj:
                    errores.append(f'cobrador "{collector_username}" no existe en tu equipo')

                if not subcartera_nombre:
                    errores.append('subcartera: vacía')

                if errores:
                    return None, errores
                return {
                    'lead': lead, 'collector': collector_obj,
                    'cartera': cartera, 'subcartera_nombre': subcartera_nombre,
                }, []

            resultado = procesar_carga(
                upload_form.cleaned_data['file'], ASSIGNMENT_UPLOAD_ALIASES,
                ('cartera', 'subcartera', 'op', 'collector'), validar_fila,
                nombre_archivo='errores_asignacion.xlsx',
            )
            if not resultado.ok:
                return resultado.respuesta_error

            with transaction.atomic():
                for f in resultado.filas:
                    subcartera = Subcartera.objects.filter(
                        cartera=f['cartera'], nombre__iexact=f['subcartera_nombre']
                    ).first()
                    if subcartera is None:
                        subcartera = Subcartera.objects.create(cartera=f['cartera'], nombre=f['subcartera_nombre'])

                    lead = f['lead']
                    lead.subcartera = subcartera
                    lead.assigned_to = f['collector']
                    if lead.activo == Lead.DESASIGNADO:
                        lead.activo = Lead.ACTIVO
                        lead.desasignado_at = None
                    lead.save()
                    LeadAssignment.objects.create(
                        lead=lead,
                        user=f['collector'],
                        assigned_by=request.user,
                    )

            messages.success(request, f"{len(resultado.filas)} lead(s) asignado(s) correctamente.")
            return redirect('leads:list')

        return render(request, self.template_name, {'upload_form': upload_form})