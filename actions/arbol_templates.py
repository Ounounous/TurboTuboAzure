"""
Plantillas de arbol de gestiones (Medios + Resultados) para las 3 carteras reales del negocio.

Las funciones `aplicar_galgo(cartera)` / `aplicar_tanner(cartera)` / `aplicar_nuevo_capital(cartera)`
son AUTOCONTENIDAS -- no piden Excel. Los datos de Galgo y Nuevo Capital se extrajeron tal cual de
las carteras ya cargadas y validadas en este ambiente (mismo criterio que Tanner, que ya estaba
transcrito del instructivo oficial). Las usa la asignacion de arbol desde la web (Carteras ->
detalle, solo admin/owner): 3 opciones fijas, sin subir archivo.

`importar_galgo_excel` / `importar_nuevo_capital_excel` quedan aparte, solo para los management
commands (uso puntual por CLI si en el futuro hay que reimportar un Excel nuevo de esas carteras
-- no se usan desde la web).
"""
from openpyxl import load_workbook

from .demografia_rules import desactiva_whatsapp, efecto_demografia
from .models import Medio, Resultado


def _crear_medios(cartera, medios):
    """medios: lista de (nombre, canal, es_llamada, es_inbound). Devuelve cuantos se crearon."""
    creados = 0
    for nombre, canal, es_llamada, es_inbound in medios:
        medio, created = Medio.objects.get_or_create(
            cartera=cartera, nombre=nombre,
            defaults={'canal': canal, 'es_llamada': es_llamada, 'es_inbound': es_inbound},
        )
        medio.canal = canal
        medio.es_llamada = es_llamada
        medio.es_inbound = es_inbound
        medio.permite_manual = medio.calcular_permite_manual()
        medio.save(update_fields=['canal', 'es_llamada', 'es_inbound', 'permite_manual'])
        if created:
            creados += 1
    return creados


# ===========================================================================
# GALGO -- 3 medios, 13 resultados (sin codigo ni tipo_contacto).
# ===========================================================================
GALGO_MEDIOS = [
    # (nombre, canal, es_llamada, es_inbound)
    ('EMAIL', Medio.CANAL_EMAIL, False, False),
    ('TELEFONICO', Medio.CANAL_TELEFONO, True, False),
    ('WHATSAPP', Medio.CANAL_TELEFONO, False, False),
]

GALGO_RESULTADOS = [
    # (nombre, contactabilidad, es_default, crea_compromiso, requiere_fecha_pago, efecto_pago, descarga_grabacion)
    ('COMPROMISO DE PAGO', Resultado.CON_CONTACTO, False, True, True, '', True),
    ('CONTACTO SIN FECHA', Resultado.CON_CONTACTO, False, False, False, '', True),
    ('DACION', Resultado.CON_CONTACTO, False, True, True, '', True),
    ('DESCONOCE DEUDA', Resultado.CON_CONTACTO, True, False, False, '', True),
    ('EMAIL INVALIDO', Resultado.SIN_CONTACTO, False, False, False, '', False),
    ('FONO NO CORRESPONDE', Resultado.SIN_CONTACTO, False, False, False, '', False),
    ('FUERA DE SERVICIO', Resultado.SIN_CONTACTO, False, False, False, '', False),
    ('MSJ DE CONTACTO', Resultado.SIN_CONTACTO, True, False, False, '', False),
    ('NO RESPONDE', Resultado.SIN_CONTACTO, True, False, False, '', False),
    ('PAGO / AL DIA', Resultado.CON_CONTACTO, False, False, True, Resultado.EFECTO_AL_DIA, True),
    ('PAGO / CONTENIDO', Resultado.CON_CONTACTO, True, False, True, Resultado.EFECTO_PAGANDO, True),
    ('RESPONDE MSJ / CONTACTO SIN FECHA', Resultado.CON_CONTACTO, False, False, False, '', True),
    ('SIN WHATSAPP', Resultado.SIN_CONTACTO, False, False, False, '', False),
]


def aplicar_galgo(cartera):
    medios_creados = _crear_medios(cartera, GALGO_MEDIOS)

    resultados_creados, resultados_actualizados = 0, 0
    for nombre, contactabilidad, es_default, crea_compromiso, requiere_fecha_pago, efecto_pago, descarga_grabacion in GALGO_RESULTADOS:
        resultado, created = Resultado.objects.get_or_create(cartera=cartera, nombre=nombre)
        resultado.contactabilidad = contactabilidad
        resultado.es_default = es_default
        resultado.crea_compromiso = crea_compromiso
        resultado.requiere_fecha_pago = requiere_fecha_pago
        resultado.efecto_pago = efecto_pago
        resultado.efecto_demografia = efecto_demografia(nombre)
        resultado.desactiva_whatsapp = desactiva_whatsapp(nombre)
        if created:
            resultado.descarga_grabacion = descarga_grabacion
            resultados_creados += 1
        else:
            resultados_actualizados += 1
        resultado.save()

    return {
        'medios_creados': medios_creados,
        'resultados_creados': resultados_creados,
        'resultados_actualizados': resultados_actualizados,
    }


# ===========================================================================
# TANNER -- 8 medios, 113 resultados (codigo + tipo_contacto). Transcrito del "Instructivo base
# de gestiones - Tanner Automotriz (version 11)".
# ===========================================================================
TANNER_MEDIOS = [
    # (codigo, nombre, canal, es_llamada)
    ('1', 'Manual', Medio.CANAL_TELEFONO, True),
    ('2', 'Discador', Medio.CANAL_TELEFONO, True),
    ('3', 'Terreno', Medio.CANAL_TELEFONO, False),
    ('4', 'IVR', Medio.CANAL_TELEFONO, True),
    ('5', 'SMS', Medio.CANAL_TELEFONO, False),
    ('6', 'Email', Medio.CANAL_EMAIL, False),
    ('7', 'WhatsApp', Medio.CANAL_TELEFONO, False),
    ('8', 'Bot', Medio.CANAL_TELEFONO, False),
]

# (codigo, tipo_contacto, respuesta, requiere_fecha_pago)
TANNER_RESULTADOS = [
    ('100', 'DIRECTO', 'PROMESA DE PAGO', True),
    ('101', 'DIRECTO', 'INTENCION DE PAGO', False),
    ('102', 'DIRECTO', 'YA PAGO', False),
    ('103', 'DIRECTO', 'PAGADO', False),
    ('104', 'DIRECTO', 'EN PROCESO DE DACION', False),
    ('105', 'DIRECTO', 'EN PROCESO DE RENEGOCIACION', False),
    ('106', 'DIRECTO', 'INCAUTADO/REMATADO', False),
    ('107', 'DIRECTO', 'NO ES LA FECHA QUE PACTO', False),
    ('108', 'DIRECTO', 'DICE HABER PAGADO', False),
    ('109', 'DIRECTO', 'SIN MODALIDAD DE PAGO', False),
    ('110', 'DIRECTO', 'SINIESTRO', False),
    ('111', 'DIRECTO', 'DESCONOCE DEUDA', False),
    ('112', 'DIRECTO', 'PAGA TERCERO O AVAL', False),
    ('113', 'DIRECTO', 'CESANTE', False),
    ('114', 'DIRECTO', 'ENFERMEDAD / LICENCIA MEDICA', False),
    ('115', 'DIRECTO', 'PROBLEMAS FINANCIEROS', False),
    ('116', 'DIRECTO', 'SOLICITA LLAMADO POSTERIOR', False),
    ('117', 'DIRECTO', 'NO QUIERE PAGAR', False),
    ('118', 'DIRECTO', 'NEGOCIACIÓN POR EMAIL', False),
    ('119', 'DIRECTO', 'NEGOCIACIÓN POR WHATSAPP', False),
    ('120', 'DIRECTO', 'VACACIONES', False),
    ('121', 'DIRECTO', 'TRAMITANDO SEGURO', False),
    ('122', 'DIRECTO', 'CONTINGENCIA', False),
    ('123', 'DIRECTO', 'RESPUESTA AGRESIVA', False),
    ('124', 'DIRECTO', 'OTROS', False),
    ('125', 'DIRECTO', 'INTENCION DE DACIÓN', False),
    ('126', 'DIRECTO', 'INTENCION DE RENEGOCIACIÓN', False),
    ('127', 'DIRECTO', 'CITACIÓN ENTREGADA A TITULAR', False),
    ('128', 'DIRECTO', 'AUTORIZA LLAMADO POSTERIOR', False),
    ('129', 'DIRECTO', 'PAC INTERESADO', False),
    ('130', 'DIRECTO', 'PAC CONTRATADO WEB', False),
    ('131', 'DIRECTO', 'PAC CONTRATADO FÍSICO', False),
    ('132', 'DIRECTO', 'PAGARA POR OTROS MEDIOS', False),
    ('133', 'DIRECTO', 'NO BANCARIZADO', False),
    ('134', 'DIRECTO', 'BANCO SIN CONVENIO', False),
    ('135', 'DIRECTO', 'NO INTERESADO SIN MOTIVO', False),
    ('136', 'DIRECTO', 'YA TIENE PAC', False),
    ('137', 'DIRECTO', 'INTERESADO EN VENTA DIRECTA', False),
    ('138', 'DIRECTO', 'NO INTERESADO EN VENTA DIRECTA (OTROS)', False),
    ('139', 'DIRECTO', 'REGULARIZARA POR OTRO MEDIO', False),
    ('140', 'DIRECTO', 'VEHICULO COMO HERRAMIENTA DE TRABAJO', False),
    ('141', 'DIRECTO', 'INFORMACION ENVIADA A TANNER', False),
    ('142', 'DIRECTO', 'EN LICITACIÓN CONCESIONARIOS', False),
    ('143', 'DIRECTO', 'SIN RESPUESTA CLIENTE', False),
    ('144', 'DIRECTO', 'SIN GESTION CONCESIONARIO', False),
    ('145', 'DIRECTO', 'EN PROCESO REVISION VEHICULO', False),
    ('146', 'DIRECTO', 'DISCONFORMIDAD POR PRECIO', False),
    ('147', 'DIRECTO', 'SIN RESPUESTA OFERTA', False),
    ('148', 'DIRECTO', 'SIN CONTACTO CONCESIONARIO', False),
    ('149', 'DIRECTO', 'VEHICULO EN MAL ESTADO', False),
    ('150', 'DIRECTO', 'VEHICULO CON DEMASIADO KM', False),
    ('151', 'DIRECTO', 'VEHICULO CON ANOTACIONES EN TRAMITE', False),
    ('152', 'DIRECTO', 'VEHICULO NO COMERCIAL', False),
    ('153', 'DIRECTO', 'VENDIDO', False),
    ('200', 'DIRECTO AVAL', 'PROMESA DE PAGO', True),
    ('201', 'DIRECTO AVAL', 'INTENCION DE PAGO', False),
    ('202', 'DIRECTO AVAL', 'YA PAGO', False),
    ('203', 'DIRECTO AVAL', 'PAGADO', False),
    ('204', 'DIRECTO AVAL', 'EN PROCESO DE DACION', False),
    ('205', 'DIRECTO AVAL', 'EN PROCESO DE RENEGOCIACION', False),
    ('206', 'DIRECTO AVAL', 'INCAUTADO/REMATADO', False),
    ('207', 'DIRECTO AVAL', 'NO ES LA FECHA QUE PACTO', False),
    ('208', 'DIRECTO AVAL', 'DICE HABER PAGADO', False),
    ('209', 'DIRECTO AVAL', 'SIN MODALIDAD DE PAGO', False),
    ('210', 'DIRECTO AVAL', 'SINIESTRO', False),
    ('211', 'DIRECTO AVAL', 'DESCONOCE DEUDA', False),
    ('213', 'DIRECTO AVAL', 'CESANTE', False),
    ('214', 'DIRECTO AVAL', 'ENFERMO', False),
    ('215', 'DIRECTO AVAL', 'SIN DINERO', False),
    ('216', 'DIRECTO AVAL', 'SOLICITA LLAMADO POSTERIOR', False),
    ('217', 'DIRECTO AVAL', 'NO QUIERE PAGAR', False),
    ('218', 'DIRECTO AVAL', 'VACACIONES', False),
    ('219', 'DIRECTO AVAL', 'TRAMITANDO SEGURO', False),
    ('220', 'DIRECTO AVAL', 'CONTINGENCIA', False),
    ('221', 'DIRECTO AVAL', 'RESPUESTA AGRESIVA', False),
    ('222', 'DIRECTO AVAL', 'OTROS', False),
    ('223', 'DIRECTO AVAL', 'AUTORIZA LLAMADO POSTERIOR', False),
    ('300', 'INDIRECTO', 'FALLECIDO', False),
    ('301', 'INDIRECTO', 'PAGA TERCERO', False),
    ('302', 'INDIRECTO', 'SE DEJA RECADO CON FAMILIAR', False),
    ('303', 'INDIRECTO', 'SE DEJA RECADO CON TERCERO', False),
    ('304', 'INDIRECTO', 'TERCERO / FAMILIAR LLAMAR MAS TARDE', False),
    ('305', 'INDIRECTO', 'TERCERO NO CONOCE A DEUDOR', False),
    ('306', 'INDIRECTO', 'DIRECCIÓN CONFIRMADA CON VECINO', False),
    ('307', 'INDIRECTO', 'CITACIÓN RECIBIDA POR TERCERO', False),
    ('308', 'INDIRECTO', 'TERCERO CONFIRMA CAMBIO DE DOMICILIO', False),
    ('309', 'INDIRECTO', 'REHUSA ATENCIÓN', False),
    ('310', 'INDIRECTO', 'AUTORIZA LLAMADO POSTERIOR', False),
    ('400', 'SIN CONTACTO OPERADOR', 'SIN CONTACTO BUZON DE VOZ', False),
    ('401', 'SIN CONTACTO OPERADOR', 'SIN CONTACTO NO CONTESTA', False),
    ('402', 'SIN CONTACTO OPERADOR', 'TELEFONO NO CORRESPONDE', False),
    ('403', 'SIN CONTACTO OPERADOR', 'CORTA LLAMADO', False),
    ('404', 'SIN CONTACTO OPERADOR', 'BUSQUEDA DE DATOS', False),
    ('405', 'SIN CONTACTO OPERADOR', 'SIN DATOS DEMOGRAFICOS', False),
    ('500', 'SIN CONTACTO MAQUINA', 'SIN OPERADOR DISPONIBLE (DROP)', False),
    ('501', 'SIN CONTACTO MAQUINA', 'SIN CONTACTO MAQUINA', False),
    ('510', 'SIN CONTACTO TERRENO', 'DIRECCIÓN NO CORRESPONDE', False),
    ('511', 'SIN CONTACTO TERRENO', 'DIRECCIÓN INEXISTENTE', False),
    ('512', 'SIN CONTACTO TERRENO', 'DIRECCIÓN DE CONOCIDO', False),
    ('513', 'SIN CONTACTO TERRENO', 'SIN MORADORES', False),
    ('514', 'SIN CONTACTO TERRENO', 'LUGAR INACCESIBLE', False),
    ('600', 'ACCION MASIVA', 'ENVIO SMS ENTREGADO', False),
    ('601', 'ACCION MASIVA', 'ENVIO CARTA ENTREGADO', False),
    ('602', 'ACCION MASIVA', 'ENVIO EMAIL ENTREGADO', False),
    ('603', 'ACCION MASIVA', 'ENVIO WHATSAPP ENTREGADO', False),
    ('604', 'ACCION MASIVA', 'ENVIO IVR ENTREGADO', False),
    ('605', 'ACCION MASIVA', 'ENVIO SMS NO ENTREGADO', False),
    ('606', 'ACCION MASIVA', 'ENVIO CARTA NO ENTREGADO', False),
    ('607', 'ACCION MASIVA', 'ENVIO EMAIL NO ENTREGADO', False),
    ('608', 'ACCION MASIVA', 'ENVIO WHATSAPP NO ENTREGADO', False),
    ('609', 'ACCION MASIVA', 'ENVIO IVR INCOMPLETO', False),
    ('610', 'ACCION MASIVA', 'ENVIO IVR NO ENTREGADO', False),
    ('800', 'SIN GESTION', 'SIN GESTION', False),
]

TANNER_TIPOS_CONTACTO_CON_CONTACTO = {'DIRECTO', 'DIRECTO AVAL', 'INDIRECTO'}
TANNER_CODIGOS_EFECTO_PAGANDO = {'112', '301'}  # PAGA TERCERO O AVAL (DIRECTO), PAGA TERCERO (INDIRECTO)
# PAC / Venta Directa (cod. 129-153): campaña de venta de vehiculo, no gestion de cobranza -- la
# paleta oficial de Tanner los marca "NO VAN A GESTIONES USUARIOS". Se sacan del formulario manual
# pero se siguen aceptando en la carga masiva/Excel (son de backoffice, no de gestor).
TANNER_CODIGOS_NO_MANUAL = {str(c) for c in range(129, 154)}


def aplicar_tanner(cartera):
    medios_creados = 0
    for codigo, nombre, canal, es_llamada in TANNER_MEDIOS:
        medio, created = Medio.objects.get_or_create(
            cartera=cartera, nombre=nombre,
            defaults={'canal': canal, 'es_llamada': es_llamada, 'codigo': codigo},
        )
        medio.codigo = codigo
        medio.canal = canal
        medio.es_llamada = es_llamada
        medio.permite_manual = medio.calcular_permite_manual()
        medio.save(update_fields=['codigo', 'canal', 'es_llamada', 'permite_manual'])
        if created:
            medios_creados += 1

    resultados_creados, resultados_actualizados = 0, 0
    for codigo, tipo_contacto, respuesta, requiere_fecha_pago in TANNER_RESULTADOS:
        con_contacto = tipo_contacto in TANNER_TIPOS_CONTACTO_CON_CONTACTO
        resultado, created = Resultado.objects.get_or_create(
            cartera=cartera, nombre=respuesta, tipo_contacto=tipo_contacto,
        )
        resultado.codigo = codigo
        resultado.contactabilidad = Resultado.CON_CONTACTO if con_contacto else Resultado.SIN_CONTACTO
        resultado.crea_compromiso = requiere_fecha_pago
        resultado.requiere_fecha_pago = requiere_fecha_pago
        resultado.efecto_pago = Resultado.EFECTO_PAGANDO if codigo in TANNER_CODIGOS_EFECTO_PAGANDO else ''
        resultado.permite_manual = codigo not in TANNER_CODIGOS_NO_MANUAL
        resultado.efecto_demografia = efecto_demografia(respuesta)
        resultado.desactiva_whatsapp = desactiva_whatsapp(respuesta)
        if created:
            resultado.descarga_grabacion = con_contacto
            resultados_creados += 1
        else:
            resultados_actualizados += 1
        resultado.save()

    return {
        'medios_creados': medios_creados,
        'resultados_creados': resultados_creados,
        'resultados_actualizados': resultados_actualizados,
    }


# ===========================================================================
# NUEVO CAPITAL -- 12 medios, 76 resultados (tipo_contacto, sin codigo).
# ===========================================================================
NC_MEDIOS = [
    # (nombre, canal, es_llamada, es_inbound)
    ('BUSQUEDA', Medio.CANAL_TELEFONO, False, False),
    ('CORREO', Medio.CANAL_EMAIL, False, False),
    ('CORREO RECIBIDO', Medio.CANAL_EMAIL, False, True),
    ('IVR', Medio.CANAL_TELEFONO, True, False),
    ('IVR AUDIO', Medio.CANAL_TELEFONO, True, False),
    ('LLAMADA', Medio.CANAL_TELEFONO, True, False),
    ('LLAMADA RECIBIDA', Medio.CANAL_TELEFONO, True, True),
    ('SMS', Medio.CANAL_TELEFONO, False, False),
    ('SMS RECIBIDO', Medio.CANAL_TELEFONO, False, True),
    ('TERRENO', Medio.CANAL_TELEFONO, False, False),
    ('WHATSAPP', Medio.CANAL_TELEFONO, False, False),
    ('WHATSAPP RECIBIDO', Medio.CANAL_TELEFONO, False, True),
]

NC_RESULTADOS = [
    # (nombre, tipo_contacto, contactabilidad, es_default, crea_compromiso, requiere_fecha_pago, efecto_pago, descarga_grabacion)
    ('BUZON DE VOZ', 'ACCION MASIVA', 'sin_contacto', False, False, False, '', False),
    ('ENVIO EMAIL ENTREGADO', 'ACCION MASIVA', 'sin_contacto', False, False, False, '', False),
    ('ENVIO EMAIL NO ENTREGADO', 'ACCION MASIVA', 'sin_contacto', False, False, False, '', False),
    ('ENVIO SMS ENTREGADO', 'ACCION MASIVA', 'sin_contacto', False, False, False, '', False),
    ('ENVIO SMS NO ENTREGADO', 'ACCION MASIVA', 'sin_contacto', False, False, False, '', False),
    ('ENVIO WHATSAPP ENTREGADO', 'ACCION MASIVA', 'sin_contacto', False, False, False, '', False),
    ('ENVIO WHATSAPP NO ENTREGADO', 'ACCION MASIVA', 'sin_contacto', False, False, False, '', False),
    ('ERROR DE CONEXION', 'ACCION MASIVA', 'sin_contacto', False, False, False, '', False),
    ('FONO NO DISPONIBLE', 'ACCION MASIVA', 'sin_contacto', False, False, False, '', False),
    ('IVR ENVIADO', 'ACCION MASIVA', 'sin_contacto', False, False, False, '', False),
    ('NO CONTESTA', 'ACCION MASIVA', 'sin_contacto', False, False, False, '', False),
    ('TONO OCUPADO', 'ACCION MASIVA', 'sin_contacto', False, False, False, '', False),
    ('UTIL POSITIVO', 'ACCION MASIVA', 'sin_contacto', False, False, False, '', False),
    ('CESANTE', 'DIRECTO', 'con_contacto', False, False, False, '', True),
    ('COMPROMISO DE PAGO', 'DIRECTO', 'con_contacto', False, True, True, '', True),
    ('DESCONOCE DEUDA', 'DIRECTO', 'con_contacto', False, False, False, '', True),
    ('DICE HABER PAGADO', 'DIRECTO', 'con_contacto', False, False, False, '', True),
    ('DOMICILIO CORRESPONDE', 'DIRECTO', 'con_contacto', False, False, False, '', False),
    ('ENFERMO', 'DIRECTO', 'con_contacto', False, False, False, '', True),
    ('EN NEGOCIACION', 'DIRECTO', 'con_contacto', False, False, False, '', True),
    ('EN PROCESO DE DACION', 'DIRECTO', 'con_contacto', False, False, False, '', True),
    ('EN PROCESO DE RENEGOCIACION', 'DIRECTO', 'con_contacto', False, False, False, '', True),
    ('ENVIO EMAIL ENTREGADO', 'DIRECTO', 'con_contacto', False, False, False, '', False),
    ('ENVIO SMS ENTREGADO', 'DIRECTO', 'con_contacto', False, False, False, '', False),
    ('ENVIO WHATSAPP ENTREGADO', 'DIRECTO', 'con_contacto', False, False, False, '', False),
    ('INCAUTADO/REMATADO', 'DIRECTO', 'con_contacto', False, False, False, '', True),
    ('INTENCION DE DACION', 'DIRECTO', 'con_contacto', False, False, False, '', True),
    ('INTENCION DE PAGO', 'DIRECTO', 'con_contacto', False, True, True, '', True),
    ('INTENCION DE RENEGOCIACION', 'DIRECTO', 'con_contacto', False, False, False, '', True),
    ('NO CORRESPONDE A TITULAR', 'DIRECTO', 'con_contacto', False, False, False, '', False),
    ('OTROS', 'DIRECTO', 'con_contacto', False, False, False, '', True),
    ('PAGA TERCERO O AVAL', 'DIRECTO', 'con_contacto', False, False, False, 'pagando', True),
    ('SIN DINERO', 'DIRECTO', 'con_contacto', False, False, False, '', True),
    ('SINIESTRO', 'DIRECTO', 'con_contacto', False, False, False, '', True),
    ('SIN INTENCION DE PAGO', 'DIRECTO', 'con_contacto', False, False, False, '', True),
    ('SMS RECIBIDO', 'DIRECTO', 'con_contacto', False, False, False, '', False),
    ('CESANTE', 'DIRECTO AVAL', 'con_contacto', False, False, False, '', True),
    ('COMPROMISO DE PAGO', 'DIRECTO AVAL', 'con_contacto', False, True, True, '', True),
    ('DESCONOCE DEUDA', 'DIRECTO AVAL', 'con_contacto', False, False, False, '', True),
    ('DICE HABER PAGADO', 'DIRECTO AVAL', 'con_contacto', False, False, False, '', True),
    ('ENFERMO', 'DIRECTO AVAL', 'con_contacto', False, False, False, '', True),
    ('EN NEGOCIACION', 'DIRECTO AVAL', 'con_contacto', False, False, False, '', True),
    ('EN PROCESO DE DACION', 'DIRECTO AVAL', 'con_contacto', False, False, False, '', True),
    ('EN PROCESO DE RENEGOCIACION', 'DIRECTO AVAL', 'con_contacto', False, False, False, '', True),
    ('ENVIO EMAIL ENTREGADO', 'DIRECTO AVAL', 'con_contacto', False, False, False, '', False),
    ('ENVIO WHATSAPP ENTREGADO', 'DIRECTO AVAL', 'con_contacto', False, False, False, '', False),
    ('INCAUTADO/REMATADO', 'DIRECTO AVAL', 'con_contacto', False, False, False, '', True),
    ('INTENCION DE DACION', 'DIRECTO AVAL', 'con_contacto', False, False, False, '', True),
    ('INTENCION DE PAGO', 'DIRECTO AVAL', 'con_contacto', False, True, True, '', True),
    ('INTENCION DE RENEGOCIACION', 'DIRECTO AVAL', 'con_contacto', False, False, False, '', True),
    ('NO CORRESPONDE A TITULAR', 'DIRECTO AVAL', 'con_contacto', False, False, False, '', False),
    ('OTROS', 'DIRECTO AVAL', 'con_contacto', False, False, False, '', True),
    ('PAGA TERCERO O AVAL', 'DIRECTO AVAL', 'con_contacto', False, False, False, 'pagando', True),
    ('SIN DINERO', 'DIRECTO AVAL', 'con_contacto', False, False, False, '', True),
    ('SINIESTRO', 'DIRECTO AVAL', 'con_contacto', False, False, False, '', True),
    ('SIN INTENCION DE PAGO', 'DIRECTO AVAL', 'con_contacto', False, False, False, '', True),
    ('EN NEGOCIACION', 'INDIRECTO', 'con_contacto', False, False, False, '', True),
    ('ENVIO EMAIL ENTREGADO', 'INDIRECTO', 'con_contacto', False, False, False, '', False),
    ('ENVIO SMS ENTREGADO', 'INDIRECTO', 'con_contacto', False, False, False, '', False),
    ('ENVIO WHATSAPP ENTREGADO', 'INDIRECTO', 'con_contacto', False, False, False, '', False),
    ('FALLECIDO', 'INDIRECTO', 'con_contacto', False, False, False, '', True),
    ('NO CORRESPONDE A TITULAR', 'INDIRECTO', 'con_contacto', False, False, False, '', True),
    ('RECADO CON TERCERO', 'INDIRECTO', 'con_contacto', False, False, False, '', True),
    ('SINIESTRO', 'INDIRECTO', 'con_contacto', False, False, False, '', True),
    ('SMS RECIBIDO', 'INDIRECTO', 'con_contacto', False, False, False, '', False),
    ('BUSQUEDA DE DATOS', 'SIN CONTACTO', 'sin_contacto', False, False, False, '', False),
    ('BUZON DE VOZ', 'SIN CONTACTO', 'sin_contacto', False, False, False, '', False),
    ('CORTA LLAMADO', 'SIN CONTACTO', 'sin_contacto', False, False, False, '', False),
    ('ENVIO EMAIL ENTREGADO', 'SIN CONTACTO', 'sin_contacto', False, False, False, '', False),
    ('ENVIO EMAIL NO ENTREGADO', 'SIN CONTACTO', 'sin_contacto', False, False, False, '', False),
    ('ENVIO WHATSAPP ENTREGADO', 'SIN CONTACTO', 'sin_contacto', False, False, False, '', False),
    ('ENVIO WHATSAPP NO ENTREGADO', 'SIN CONTACTO', 'sin_contacto', False, False, False, '', False),
    ('FUERA DE SERVICIO', 'SIN CONTACTO', 'sin_contacto', False, False, False, '', False),
    ('NO CONTESTA', 'SIN CONTACTO', 'sin_contacto', False, False, False, '', False),
    ('SIN DATOS DEMOGRAFICOS', 'SIN CONTACTO', 'sin_contacto', False, False, False, '', False),
    ('TONO OCUPADO', 'SIN CONTACTO', 'sin_contacto', False, False, False, '', False),
]


def aplicar_nuevo_capital(cartera):
    medios_creados = _crear_medios(cartera, NC_MEDIOS)

    resultados_creados, resultados_actualizados = 0, 0
    for nombre, tipo_contacto, contactabilidad, es_default, crea_compromiso, requiere_fecha_pago, efecto_pago, descarga_grabacion in NC_RESULTADOS:
        resultado, created = Resultado.objects.get_or_create(cartera=cartera, nombre=nombre, tipo_contacto=tipo_contacto)
        resultado.contactabilidad = contactabilidad
        resultado.es_default = es_default
        resultado.crea_compromiso = crea_compromiso
        resultado.requiere_fecha_pago = requiere_fecha_pago
        resultado.efecto_pago = efecto_pago
        resultado.efecto_demografia = efecto_demografia(nombre)
        resultado.desactiva_whatsapp = desactiva_whatsapp(nombre)
        if created:
            resultado.descarga_grabacion = descarga_grabacion
            resultados_creados += 1
        else:
            resultados_actualizados += 1
        resultado.save()

    return {
        'medios_creados': medios_creados,
        'resultados_creados': resultados_creados,
        'resultados_actualizados': resultados_actualizados,
    }


APLICAR_POR_TIPO = {
    'galgo': aplicar_galgo,
    'tanner': aplicar_tanner,
    'nuevo_capital': aplicar_nuevo_capital,
}


# ===========================================================================
# Solo para los management commands (CLI): reimportar desde un Excel de origen si hiciera falta
# actualizar el arbol de Galgo/Nuevo Capital mas adelante. NO se usan desde la asignacion web.
# ===========================================================================
GALGO_CANAL_POR_MEDIO = {
    'WHATSAPP': Medio.CANAL_TELEFONO,
    'TELEFONICO': Medio.CANAL_TELEFONO,
    'SMS': Medio.CANAL_TELEFONO,
    'EMAIL': Medio.CANAL_EMAIL,
}
GALGO_MEDIOS_LLAMADA = {'TELEFONICO', 'LLAMADA', 'LLAMADA TELEFONICA', 'VOZ'}


def importar_galgo_excel(cartera, excel_file):
    wb = load_workbook(excel_file, data_only=True)
    ws = wb.active

    medios_creados = 0
    filas_omitidas = 0
    resultados_agg = {}

    for row in ws.iter_rows(min_row=3, values_only=True):
        if not row or len(row) < 2:
            continue
        medio_nombre, resultado_nombre, contactabilidad, es_default, crea_compromiso = (list(row) + [None] * 5)[:5]

        if not medio_nombre or not resultado_nombre:
            filas_omitidas += 1
            continue

        medio_nombre = str(medio_nombre).strip()
        resultado_nombre = str(resultado_nombre).strip()
        canal = GALGO_CANAL_POR_MEDIO.get(medio_nombre.upper(), Medio.CANAL_TELEFONO)

        medio, medio_created = Medio.objects.get_or_create(
            cartera=cartera, nombre=medio_nombre,
            defaults={'canal': canal, 'es_llamada': medio_nombre.upper() in GALGO_MEDIOS_LLAMADA}
        )
        nuevo_permite = medio.calcular_permite_manual()
        if medio.permite_manual != nuevo_permite:
            medio.permite_manual = nuevo_permite
            medio.save(update_fields=['permite_manual'])
        if medio_created:
            medios_creados += 1

        contactabilidad_norm = (
            Resultado.CON_CONTACTO
            if 'CON CONTACTO' in str(contactabilidad or '').upper()
            else Resultado.SIN_CONTACTO
        )
        crea_compromiso_bool = bool(crea_compromiso)

        agg = resultados_agg.setdefault(resultado_nombre, {
            'contactabilidad': contactabilidad_norm,
            'es_default': False,
            'crea_compromiso': False,
            'es_llamada': False,
        })
        agg['es_default'] = agg['es_default'] or bool(es_default)
        agg['crea_compromiso'] = agg['crea_compromiso'] or crea_compromiso_bool
        agg['es_llamada'] = agg['es_llamada'] or medio.es_llamada
        if contactabilidad_norm == Resultado.CON_CONTACTO:
            agg['contactabilidad'] = Resultado.CON_CONTACTO

    resultados_creados, resultados_actualizados = 0, 0
    for nombre, agg in resultados_agg.items():
        requiere_fecha_pago = agg['crea_compromiso'] or 'PAGO' in nombre.upper()
        nombre_upper = nombre.upper()
        if nombre_upper.startswith('PAGO') and 'AL DIA' in nombre_upper:
            efecto_pago = Resultado.EFECTO_AL_DIA
        elif nombre_upper.startswith('PAGO'):
            efecto_pago = Resultado.EFECTO_PAGANDO
        else:
            efecto_pago = ''

        resultado, resultado_created = Resultado.objects.get_or_create(cartera=cartera, nombre=nombre)
        resultado.contactabilidad = agg['contactabilidad']
        resultado.es_default = agg['es_default']
        resultado.crea_compromiso = agg['crea_compromiso']
        resultado.requiere_fecha_pago = requiere_fecha_pago
        resultado.efecto_pago = efecto_pago
        resultado.efecto_demografia = efecto_demografia(nombre)
        resultado.desactiva_whatsapp = desactiva_whatsapp(nombre)
        if resultado_created:
            resultado.descarga_grabacion = (
                agg['es_llamada'] and agg['contactabilidad'] == Resultado.CON_CONTACTO
            )
            resultados_creados += 1
        else:
            resultados_actualizados += 1
        resultado.save()

    return {
        'medios_creados': medios_creados,
        'resultados_creados': resultados_creados,
        'resultados_actualizados': resultados_actualizados,
        'filas_omitidas': filas_omitidas,
    }


NC_CANAL_EMAIL_ACCIONES = {'CORREO', 'CORREO RECIBIDO'}
NC_LLAMADA_ACCIONES = {'LLAMADA', 'LLAMADA RECIBIDA', 'IVR', 'IVR AUDIO'}
NC_TIPOS_CON_CONTACTO = {'DIRECTO', 'DIRECTO AVAL', 'INDIRECTO'}


def importar_nuevo_capital_excel(cartera, excel_file):
    wb = load_workbook(excel_file, data_only=True)
    ws = wb.active

    medios_info = {}
    resultados_info = {}

    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or len(row) < 7:
            continue
        accion, sub_estado, estado, in_out = row[2], row[3], row[4], row[6]
        if not accion or not estado:
            continue

        accion = str(accion).strip()
        sub_estado = str(sub_estado or '').strip()
        estado = str(estado).strip()
        accion_upper = accion.upper()

        es_inbound = 'RECIBID' in accion_upper or 'in' in str(in_out or '').strip().lower()
        canal = Medio.CANAL_EMAIL if accion_upper in NC_CANAL_EMAIL_ACCIONES else Medio.CANAL_TELEFONO
        es_llamada = accion_upper in NC_LLAMADA_ACCIONES

        m = medios_info.setdefault(accion, {'canal': canal, 'es_llamada': es_llamada, 'es_inbound': es_inbound})
        m['es_inbound'] = m['es_inbound'] or es_inbound
        m['es_llamada'] = m['es_llamada'] or es_llamada

        con_contacto = sub_estado.upper() in NC_TIPOS_CON_CONTACTO
        r = resultados_info.setdefault((estado, sub_estado), {'con_contacto': con_contacto, 'es_llamada': False})
        r['con_contacto'] = r['con_contacto'] or con_contacto
        r['es_llamada'] = r['es_llamada'] or es_llamada

    medios_creados = 0
    for nombre, info in medios_info.items():
        medio, created = Medio.objects.get_or_create(
            cartera=cartera, nombre=nombre,
            defaults={'canal': info['canal'], 'es_llamada': info['es_llamada'], 'es_inbound': info['es_inbound']},
        )
        medio.canal = info['canal']
        medio.es_llamada = info['es_llamada']
        medio.es_inbound = info['es_inbound']
        medio.permite_manual = medio.calcular_permite_manual()
        medio.save(update_fields=['canal', 'es_llamada', 'es_inbound', 'permite_manual'])
        if created:
            medios_creados += 1

    resultados_creados, resultados_actualizados = 0, 0
    for (nombre, tipo_contacto), info in resultados_info.items():
        con_contacto = info['con_contacto']
        nombre_up = nombre.upper()
        crea_compromiso = (
            'COMPROMISO DE PAGO' in nombre_up
            or ('INTENCION DE PAGO' in nombre_up and 'SIN INTENCION' not in nombre_up)
        )
        requiere_fecha = crea_compromiso
        efecto_pago = Resultado.EFECTO_PAGANDO if 'PAGA TERCERO' in nombre.upper() else ''

        resultado, created = Resultado.objects.get_or_create(cartera=cartera, nombre=nombre, tipo_contacto=tipo_contacto)
        resultado.contactabilidad = Resultado.CON_CONTACTO if con_contacto else Resultado.SIN_CONTACTO
        resultado.crea_compromiso = crea_compromiso
        resultado.requiere_fecha_pago = requiere_fecha
        resultado.efecto_pago = efecto_pago
        resultado.efecto_demografia = efecto_demografia(nombre)
        resultado.desactiva_whatsapp = desactiva_whatsapp(nombre)
        if created:
            resultado.descarga_grabacion = info['es_llamada'] and con_contacto
            resultados_creados += 1
        else:
            resultados_actualizados += 1
        resultado.save()

    return {
        'medios_creados': medios_creados,
        'resultados_creados': resultados_creados,
        'resultados_actualizados': resultados_actualizados,
    }
