import datetime

from django.contrib import messages
from django.contrib.auth.models import User
from django.core.exceptions import PermissionDenied
from django.db import transaction
from django.db.models import Q, Sum, Count, Max
from django.shortcuts import get_object_or_404, redirect,render
from django.urls import reverse
from django.utils import timezone
from django.views.generic import CreateView, DetailView, ListView
from django.contrib.auth.mixins import LoginRequiredMixin
from django.http import JsonResponse
from django.http import HttpResponse, FileResponse
from django.views import View
from django.utils.dateparse import parse_date
from openpyxl import load_workbook
from .models import Action, Medio, Resultado, PendingPbxCall, CallRecording, PaymentCommitment, Payment
from .pbx_client import PbxClient, PbxError, get_pbx_master_client
from lead.models import Lead
from lead.permissions import (
    carteras_visibles, es_admin_owner, es_supervisor, leads_visibles, scope_por_lead, subcarteras_visibles,
)
from team.models import Team
from cartera.models import Cartera, Subcartera
from demographics.models import CONTACT_ACTIVE, IDDemographics, Phone
from demographics.views import find_lead, _normalize_header
from .forms import ActionForm, AddEmailQuickForm, AddPhoneQuickForm, LeadSearchForm, DemographicSelectionForm, PaymentForm
from core.concurrency import con_limite_concurrencia
import openpyxl
import re
from io import BytesIO
import logging

logger = logging.getLogger(__name__)


def _period_bounds(today):
    """(hoy_ini, hoy_fin_excl, semana_ini, semana_fin_excl, mes_ini, mes_fin_excl)."""
    week_start, week_end, month_start, next_month = _rango_semana_mes(today)
    return (
        today, today + datetime.timedelta(days=1),
        week_start, week_end + datetime.timedelta(days=1),
        month_start, next_month,
    )


def _by_field_counts(qs, field, start, end_excl):
    # Rango con zona en vez de created_at__date__gte/__lt: mismo criterio, pero permite usar el
    # indice de Action.created_at (ver core/timeutils.py). Esta funcion corre 6 veces por carga
    # del panel de Gestiones, asi que era el recorrido de tabla mas repetido del sistema.
    from core.timeutils import rango_local
    ini, fin = rango_local(start, end_excl)
    rows = (
        qs.filter(created_at__gte=ini, created_at__lt=fin)
        .values(field).annotate(n=Count('id'))
    )
    return {(r[field] or '—'): r['n'] for r in rows}


class ActionIndexView(LoginRequiredMixin, View):
    """
    Panel de Gestiones: comparación Día/Semana/Mes por usuario y por resultado, resumen de
    compromisos y pagos, y un listado de gestiones filtrable (con las mismas acciones rápidas
    que la lista de Clientes: ver, gestionar, nota, favorito).
    """

    def get(self, request, *args, **kwargs):
        all_carteras = carteras_visibles(request.user)

        if es_supervisor(request.user):
            # Supervisor/owner/admin: todo lo de sus carteras (admin/owner: todas).
            base_actions = scope_por_lead(Action.objects.all(), request.user)
            leads_scope = leads_visibles(request.user)
            pc_scope = scope_por_lead(PaymentCommitment.objects.all(), request.user)
            pay_scope = scope_por_lead(Payment.objects.all(), request.user)
        else:
            # Cobrador: su propio trabajo.
            base_actions = Action.objects.filter(user=request.user)
            leads_scope = leads_visibles(request.user)
            pc_scope = PaymentCommitment.objects.filter(created_by=request.user)
            pay_scope = Payment.objects.filter(created_by=request.user)

        today = timezone.localdate()
        hoy_ini, hoy_fin, sem_ini, sem_fin, mes_ini, mes_fin = _period_bounds(today)

        # ---- Gestiones por usuario (Día/Semana/Mes) ----
        by_user_hoy = _by_field_counts(base_actions, 'user__username', hoy_ini, hoy_fin)
        by_user_sem = _by_field_counts(base_actions, 'user__username', sem_ini, sem_fin)
        by_user_mes = _by_field_counts(base_actions, 'user__username', mes_ini, mes_fin)
        from core.timeutils import rango_local, rango_del_dia
        mes_ini_dt, mes_fin_dt = rango_local(mes_ini, mes_fin)
        contact_rows = base_actions.filter(created_at__gte=mes_ini_dt, created_at__lt=mes_fin_dt).values(
            'user__username'
        ).annotate(total=Count('id'), con=Count('id', filter=Q(resultado__contactabilidad='con_contacto')))
        contact_by_user = {
            (r['user__username'] or '—'): (round(100 * r['con'] / r['total']) if r['total'] else 0)
            for r in contact_rows
        }
        usernames = sorted(
            set(by_user_hoy) | set(by_user_sem) | set(by_user_mes),
            key=lambda u: -by_user_mes.get(u, 0),
        )
        usuarios_tabla = [
            {
                'user': u, 'hoy': by_user_hoy.get(u, 0), 'semana': by_user_sem.get(u, 0),
                'mes': by_user_mes.get(u, 0), 'contact': contact_by_user.get(u, 0),
            }
            for u in usernames
        ]

        # ---- Resultados (Día/Semana/Mes) ----
        by_res_hoy = _by_field_counts(base_actions, 'resultado__nombre', hoy_ini, hoy_fin)
        by_res_sem = _by_field_counts(base_actions, 'resultado__nombre', sem_ini, sem_fin)
        by_res_mes = _by_field_counts(base_actions, 'resultado__nombre', mes_ini, mes_fin)
        resultado_names = sorted(
            set(by_res_hoy) | set(by_res_sem) | set(by_res_mes),
            key=lambda r: -by_res_mes.get(r, 0),
        )[:8]
        resultados_tabla = [
            {'resultado': r, 'hoy': by_res_hoy.get(r, 0), 'semana': by_res_sem.get(r, 0), 'mes': by_res_mes.get(r, 0)}
            for r in resultado_names
        ]

        # ---- Compromisos y pagos (Día/Semana/Mes) ----
        compromisos_periodo = {
            'hoy': _resumen(pc_scope.filter(fecha_compromiso=today)),
            'semana': _resumen(pc_scope.filter(fecha_compromiso__gte=sem_ini, fecha_compromiso__lt=sem_fin)),
            'mes': _resumen(pc_scope.filter(fecha_compromiso__gte=mes_ini, fecha_compromiso__lt=mes_fin)),
        }
        pagos_periodo = {
            'hoy': _resumen(pay_scope.filter(fecha=today)),
            'semana': _resumen(pay_scope.filter(fecha__gte=sem_ini, fecha__lt=sem_fin)),
            'mes': _resumen(pay_scope.filter(fecha__gte=mes_ini, fecha__lt=mes_fin)),
        }

        # ---- KPIs (sin "mejor ejecutivo") ----
        hoy_ini_dt, hoy_fin_dt = rango_del_dia(today)
        gestiones_hoy_total = base_actions.filter(
            created_at__gte=hoy_ini_dt, created_at__lt=hoy_fin_dt
        ).count()
        gestiones_hoy_con = base_actions.filter(
            created_at__gte=hoy_ini_dt, created_at__lt=hoy_fin_dt,
            resultado__contactabilidad='con_contacto',
        ).count()
        contactabilidad_hoy = round(100 * gestiones_hoy_con / gestiones_hoy_total) if gestiones_hoy_total else 0
        cutoff = today - datetime.timedelta(days=7)
        sin_gestionar = leads_scope.annotate(last_action=Max('actions__created_at')).filter(
            Q(last_action__date__lt=cutoff) | Q(last_action__isnull=True)
        ).count()

        # ---- Listado de gestiones (filtrable, con acciones rápidas por cliente) ----
        f_periodo = request.GET.get('f_periodo', 'hoy')
        if f_periodo not in ('hoy', 'semana', 'mes'):
            f_periodo = 'hoy'
        f_usuario = request.GET.get('f_usuario', '').strip()
        f_resultado = request.GET.get('f_resultado', '').strip()
        f_cartera = request.GET.get('f_cartera', '').strip()
        f_solo_compromiso = request.GET.get('f_solo_compromiso', '') == '1'

        listado = base_actions.select_related('lead__subcartera__cartera', 'medio', 'resultado', 'user')
        if f_periodo == 'semana':
            ini_dt, fin_dt = rango_local(sem_ini, sem_fin)
        elif f_periodo == 'mes':
            ini_dt, fin_dt = mes_ini_dt, mes_fin_dt
        else:
            ini_dt, fin_dt = hoy_ini_dt, hoy_fin_dt
        listado = listado.filter(created_at__gte=ini_dt, created_at__lt=fin_dt)
        if f_usuario:
            listado = listado.filter(user__username=f_usuario)
        if f_resultado:
            listado = listado.filter(resultado__nombre=f_resultado)
        if f_cartera:
            listado = listado.filter(subcartera__cartera_id=f_cartera)
        if f_solo_compromiso:
            listado = listado.filter(create_payment_commitment=True)
        listado = listado.order_by('-created_at')

        # "Mostrar: 10/50/100/Todos" -- mismo patron que Compromisos de pago.
        limit_raw = request.GET.get('limit', '10')
        if limit_raw == 'todos':
            limit = None
        else:
            try:
                limit = int(limit_raw)
            except ValueError:
                limit = 10
            if limit not in (10, 50, 100):
                limit = 10
        listado = list(listado if limit is None else listado[:limit])

        def _url_limit(value):
            params = request.GET.copy()
            params['limit'] = value
            return '?' + params.urlencode()

        fav_ids = set(
            Lead.objects.filter(
                favorited_by=request.user, pk__in=[a.lead_id for a in listado]
            ).values_list('pk', flat=True)
        )
        for a in listado:
            a.lead.is_fav = a.lead_id in fav_ids

        # Subcarteras de Tanner que el usuario supervisa (para el boton "por subcartera" del
        # reporte Tanner). Vacio para casi todo el mundo -- solo aplica si Tanner existe y el
        # usuario tiene 1+ subcartera propia ahi.
        tanner_subcarteras = subcarteras_visibles(request.user).filter(
            cartera__nombre__iexact='Tanner'
        ).select_related('cartera')

        # Equipos para "Descargar gestiones del equipo": admin/owner puede bajar cualquiera;
        # supervisor/cobrador solo los suyos (coincide exacto con lo que ya permite
        # ActionDownloadExcelView -- asi el desplegable nunca ofrece un equipo que despues de
        # un 403 al intentar bajarlo).
        equipos_descarga = Team.objects.all() if es_admin_owner(request.user) else request.user.teams.all()

        return render(
            request,
            'actions/action_index.html',
            {
                'all_teams': equipos_descarga,
                'all_carteras': all_carteras,
                'tanner_subcarteras': tanner_subcarteras,
                'usuarios_tabla': usuarios_tabla,
                'resultados_tabla': resultados_tabla,
                'compromisos_periodo': compromisos_periodo,
                'pagos_periodo': pagos_periodo,
                'gestiones_hoy_total': gestiones_hoy_total,
                'contactabilidad_hoy': contactabilidad_hoy,
                'sin_gestionar': sin_gestionar,
                'listado': listado,
                'f_periodo': f_periodo,
                'f_usuario': f_usuario,
                'f_resultado': f_resultado,
                'f_cartera': f_cartera,
                'f_solo_compromiso': f_solo_compromiso,
                'usuarios_choices': usernames,
                'resultados_choices': sorted(set(by_res_hoy) | set(by_res_sem) | set(by_res_mes)),
                'limit': 'todos' if limit is None else limit,
                'url_limit_10': _url_limit(10),
                'url_limit_50': _url_limit(50),
                'url_limit_100': _url_limit(100),
                'url_limit_todos': _url_limit('todos'),
            }
        )

class ActionCreateView(LoginRequiredMixin, CreateView):
    model = Action
    form_class = ActionForm
    template_name = 'actions/action_form.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['lead'] = self.get_lead()
        return context

    def get_lead(self):
        return get_object_or_404(Lead, pk=self.kwargs['lead_id'])

    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs['lead'] = self.get_lead()
        return kwargs

    def get_success_url(self):
        return reverse('leads:detail', kwargs={'pk': self.kwargs['lead_id']})

    def form_valid(self, form):
        lead = self.get_lead()
        if not lead.es_gestionable:
            messages.error(self.request, lead.motivo_no_gestionable)
            return redirect('leads:detail', pk=lead.pk)
        form.instance.lead = lead
        logger.debug("Form is valid. Saving the action.")
        return super().form_valid(form)

    def form_invalid(self, form):
        # Log form errors for debugging
        logger.debug(f"Form is invalid. Errors: {form.errors}")
        return super().form_invalid(form)


class ActionDetailView(LoginRequiredMixin, DetailView):
    model = Action
    template_name = 'actions/action_detail.html'
    context_object_name = 'action'

class MultiStepActionView(LoginRequiredMixin, View):
    def _get_lead(self, request, lead_id):
        """Trae el lead SOLO si esta dentro del alcance del usuario (cobrador: suyos; supervisor:
        sus carteras; admin/owner: todo). Fuera de alcance -> 404, no se distingue de inexistente.
        Sin esto, cualquiera podia ver/gestionar un lead ajeno pasando su id en la URL/sesion."""
        return get_object_or_404(leads_visibles(request.user), id=lead_id)

    def _blocked_response(self, request, lead, step):
        """Un lead suspendido/terminado no admite gestiones (si notas). Muestra el motivo."""
        return render(request, 'actions/multistep_form.html', {
            'lead': lead, 'step': step, 'no_gestionable': lead.motivo_no_gestionable,
        })

    def _step2_context(self, lead, demographic_form=None, quick_phone_form=None, quick_email_form=None):
        idd = IDDemographics.objects.filter(lead=lead).first()
        return {
            'lead': lead,
            'step': 2,
            'demographic_form': demographic_form or DemographicSelectionForm(lead=lead),
            'quick_phone_form': quick_phone_form or AddPhoneQuickForm(),
            'quick_email_form': quick_email_form or AddEmailQuickForm(),
            # Se puede agregar correo siempre: el primero queda como principal; los siguientes se
            # guardan como correos adicionales (modelo Email), sin pisar el principal.
            'puede_agregar_email': True,
            'lead_notes': lead.notes.select_related('author'),
        }

    def get(self, request, step=1, lead_id=None):
        if lead_id:
            lead = self._get_lead(request, lead_id)
            request.session['selected_lead_id'] = lead.id
            return redirect('actions:multistep_step', step=2)

        lead_id = request.session.get('selected_lead_id')

        if step == 1:
            form = LeadSearchForm()
            return render(request, 'actions/multistep_form.html', {
                'form': form, 'step': step
            })
        elif step == 2 and lead_id:
            lead = self._get_lead(request, lead_id)
            if not lead.es_gestionable:
                return self._blocked_response(request, lead, step)
            return render(request, 'actions/multistep_form.html', self._step2_context(lead))
        elif step == 3 and lead_id:
            lead = self._get_lead(request, lead_id)
            if not lead.es_gestionable:
                return self._blocked_response(request, lead, step)
            selected_phone_id = request.session.get('selected_phone')
            selected_email = request.session.get('selected_email')
            canal = Medio.CANAL_TELEFONO if selected_phone_id else Medio.CANAL_EMAIL
            selected_phone = Phone.objects.filter(pk=selected_phone_id).first()
            userprofile = getattr(request.user, 'userprofile', None)
            es_cobrador = bool(userprofile and userprofile.user_type == 'collector')
            form = ActionForm(cartera=lead.subcartera.cartera, canal=canal, es_cobrador=es_cobrador)
            from .gestion_defaults import resultados_default_por_medio
            resultado_defaults = resultados_default_por_medio(lead.subcartera.cartera, lead=lead)
            return render(request, 'actions/multistep_form.html', {
                'lead': lead, 'action_form': form, 'step': step,
                'canal': canal,
                'selected_phone': selected_phone,
                'phone_digits': re.sub(r'\D', '', selected_phone.phone_number) if selected_phone else '',
                'selected_email': selected_email,
                'has_pbx': bool(userprofile and userprofile.has_pbx_credentials),
                'lead_notes': lead.notes.select_related('author'),
                'resultado_defaults': resultado_defaults,
            })
        else:
            return redirect('actions:multistep_step', step=1)

    def post(self, request, step):
        lead_id = request.session.get('selected_lead_id')
        lead = self._get_lead(request, lead_id)
        if not lead.es_gestionable:
            return self._blocked_response(request, lead, step)
        if step == 2:
            accion = request.POST.get('accion', 'seleccionar')

            if accion == 'agregar_telefono':
                quick_form = AddPhoneQuickForm(request.POST)
                if quick_form.is_valid():
                    phone = Phone.objects.create(
                        lead=lead, phone_number=quick_form.cleaned_data['phone_number'],
                        phone_type=Phone.PRINCIPAL, phone_number_status=Phone.ACTIVE,
                    )
                    request.session['selected_phone'] = phone.id
                    request.session['selected_email'] = None
                    return redirect('actions:multistep_step', step=3)
                return render(request, 'actions/multistep_form.html', self._step2_context(lead, quick_phone_form=quick_form))

            if accion == 'agregar_email':
                quick_form = AddEmailQuickForm(request.POST)
                if quick_form.is_valid():
                    nuevo = quick_form.cleaned_data['email']
                    idd = IDDemographics.objects.filter(lead=lead).first()
                    if not (idd and idd.principal_email):
                        # Primer correo del lead -> queda como principal (integra con reportes,
                        # carga masiva y "Estado de correos").
                        idd = idd or IDDemographics.objects.create(lead=lead)
                        idd.principal_email = nuevo
                        idd.principal_email_status = CONTACT_ACTIVE
                        idd.save(update_fields=['principal_email', 'principal_email_status'])
                    elif nuevo != idd.principal_email:
                        # Ya hay principal -> se guarda como correo adicional (sin duplicar).
                        from demographics.models import Email
                        Email.objects.get_or_create(lead=lead, email=nuevo)
                    request.session['selected_email'] = nuevo
                    request.session['selected_phone'] = None
                    return redirect('actions:multistep_step', step=3)
                return render(request, 'actions/multistep_form.html', self._step2_context(lead, quick_email_form=quick_form))

            form = DemographicSelectionForm(request.POST, lead=lead)
            if form.is_valid():
                selected_phone = form.cleaned_data['phone']
                selected_email = form.cleaned_data['email']
                request.session['selected_phone'] = selected_phone.id if selected_phone else None
                request.session['selected_email'] = selected_email
                return redirect('actions:multistep_step', step=3)
            return render(request, 'actions/multistep_form.html', self._step2_context(lead, demographic_form=form))
        elif step == 3:
            selected_phone_id = request.session.get('selected_phone')
            selected_email = request.session.get('selected_email')
            canal = Medio.CANAL_TELEFONO if selected_phone_id else Medio.CANAL_EMAIL
            userprofile = getattr(request.user, 'userprofile', None)
            es_cobrador = bool(userprofile and userprofile.user_type == 'collector')

            # es_cobrador tambien aca (no solo en el GET): sin esto, un cobrador podria mandar
            # por POST un resultado fuera de la lista permitida (Indirecto/Accion masiva/etc.)
            # y el queryset sin filtrar lo habria aceptado igual.
            form = ActionForm(request.POST, cartera=lead.subcartera.cartera, canal=canal, es_cobrador=es_cobrador)

            if form.is_valid():
                action = form.save(commit=False)
                action.lead = lead
                action.user = request.user  # Ensure the user is set

                # Ensure previously collected data is set
                action.phone_id = selected_phone_id  # ForeignKey to Phone model
                action.email = selected_email if not selected_phone_id else None  # Only save email if phone is not selected
                action.save()

                if selected_phone_id and action.medio.es_llamada:
                    # TODAS las llamadas sin resolver a este telefono, no solo la ultima: si el
                    # colector llamo mas de una vez antes de guardar la gestion (ej. no contesto
                    # y volvio a intentar), esta gestion describe ambos intentos por igual -- si
                    # solo se linkeara la mas reciente, las anteriores quedarian huerfanas para
                    # siempre (sin accion asociada, nunca se les busca grabacion).
                    PendingPbxCall.objects.filter(
                        user=request.user, lead=lead, phone_id=selected_phone_id, resolved=False,
                    ).update(action=action)

                logger.debug(f"Action saved: {action}, Lead: {lead}, User: {request.user}")
                messages.success(request, 'Gestión guardada correctamente.')

                if request.POST.get('submit_action') == 'add_another':
                    # Keep the lead selected, forget the phone/email so the user picks again
                    request.session.pop('selected_phone', None)
                    request.session.pop('selected_email', None)
                    return redirect('actions:multistep_step', step=2)

                request.session.pop('selected_lead_id', None)
                request.session.pop('selected_phone', None)
                request.session.pop('selected_email', None)

                return redirect('actions:popup_close')
            else:
                logger.debug(f"Invalid form data: {form.errors}")

            selected_phone = Phone.objects.filter(pk=selected_phone_id).first()
            from .gestion_defaults import resultados_default_por_medio
            return render(request, 'actions/multistep_form.html', {
                'lead': lead,
                'action_form': form,
                'step': step,
                'canal': canal,
                'selected_phone': selected_phone,
                'phone_digits': re.sub(r'\D', '', selected_phone.phone_number) if selected_phone else '',
                'selected_email': selected_email,
                'resultado_defaults': resultados_default_por_medio(lead.subcartera.cartera, lead=lead),
            })

        return redirect('actions:multistep_step', step=1)


class CancelActionView(LoginRequiredMixin, View):
    def get(self, request, lead_id, *args, **kwargs):
        for key in ('selected_lead_id', 'selected_phone', 'selected_email'):
            request.session.pop(key, None)
        return redirect('actions:popup_close')


class PopupCloseView(LoginRequiredMixin, View):
    def get(self, request, *args, **kwargs):
        return render(request, 'actions/popup_close.html')


class OriginatePbxCallView(LoginRequiredMixin, View):
    """Originates a real call through pbxip.cl using the current user's own credentials."""

    def post(self, request, *args, **kwargs):
        userprofile = getattr(request.user, 'userprofile', None)
        if not userprofile:
            return JsonResponse({'ok': False, 'reason': 'no_credentials'}, status=400)

        # Modo maestro (una cuenta admin origina/consulta): el usuario solo necesita su extension.
        # Modo por-usuario: credenciales completas. En ambos, sin extension no se puede llamar ni
        # cruzar la grabacion despues -> se avisa.
        master = get_pbx_master_client()
        if master is not None:
            if not userprofile.pbx_extension:
                return JsonResponse({'ok': False, 'reason': 'no_extension'}, status=400)
        elif not userprofile.has_pbx_credentials:
            return JsonResponse({'ok': False, 'reason': 'no_credentials'}, status=400)

        lead_id = request.POST.get('lead_id')
        phone_id = request.POST.get('phone_id')
        lead = get_object_or_404(Lead, pk=lead_id)
        if not lead.es_gestionable:
            return JsonResponse(
                {'ok': False, 'reason': 'no_gestionable', 'detail': lead.motivo_no_gestionable},
                status=403,
            )
        phone = get_object_or_404(Phone, pk=phone_id, lead=lead)
        destination = re.sub(r'\D', '', phone.phone_number)

        if not destination:
            return JsonResponse({'ok': False, 'reason': 'invalid_number'}, status=400)

        client = master if master is not None else PbxClient(userprofile.pbx_email, userprofile.pbx_password)
        try:
            client.originate_call(userprofile.pbx_extension, destination)
        except PbxError as exc:
            logger.error(f"PBX originate_call failed for user {request.user.id}: {exc}")
            return JsonResponse({'ok': False, 'reason': 'pbx_error', 'detail': str(exc)}, status=502)

        PendingPbxCall.objects.create(
            user=request.user, lead=lead, phone=phone, destination=destination,
        )

        return JsonResponse({'ok': True})

class ActionDownloadExcelView(LoginRequiredMixin, View):
    @con_limite_concurrencia('export_excel', slots=2)
    def get(self, request, *args, **kwargs):
        scope = kwargs.get('scope', 'team')

        logger.debug(f"Scope: {scope}, User: {request.user}")

        # Team ID from GET request
        team_id = request.GET.get('team_id')  # Query parameter to select the team

        # Fetch the team if 'scope' is 'team'
        if scope == 'team':
            if not team_id:  # Handle missing team_id
                return JsonResponse(
                    {"error": "No team selected. Please provide a valid team ID to download actions."},
                    status=400
                )

            try:
                # Locate the team by ID
                team = Team.objects.get(id=team_id)
                logger.debug(f"Selected team: {team} (ID: {team_id})")
            except Team.DoesNotExist:
                logger.error(f"Team with ID {team_id} does not exist.")
                return JsonResponse(
                    {"error": f"Team with ID {team_id} does not exist."},
                    status=404
                )

            # admin/owner pueden descargar cualquier equipo; el resto solo el suyo (evita que
            # un cobrador/supervisor baje gestiones de otro equipo cambiando el team_id en la URL).
            if not es_admin_owner(request.user) and not team.members.filter(pk=request.user.pk).exists():
                raise PermissionDenied

        # Handle user-specific download request (if scope == 'user')
        if scope == 'user':
            logger.debug("Downloading actions for user scope.")

        # Optional cartera filter (each cartera can pull its own report)
        cartera_id = request.GET.get('cartera')

        # Rango de fechas OBLIGATORIO: sin el, este export barria toda la historia del equipo a un
        # workbook en memoria (a 2-3 anios de datos, riesgo de OOM en el B1). Se exige start+end y
        # se acota la ventana a un maximo (por defecto 366 dias) para que un rango absurdo no lo
        # vuelva a abrir sin querer.
        start_date = request.GET.get('start_date')
        end_date = request.GET.get('end_date')
        if not start_date or not end_date:
            return JsonResponse(
                {"error": "Debes indicar el rango de fechas (start_date y end_date, YYYY-MM-DD)."},
                status=400,
            )
        try:
            start_date = parse_date(start_date)
            end_date = parse_date(end_date)
        except ValueError:
            return JsonResponse({"error": "Formato de fecha inválido. Usa YYYY-MM-DD."}, status=400)
        if not start_date or not end_date:
            return JsonResponse({"error": "Formato de fecha inválido. Usa YYYY-MM-DD."}, status=400)
        if start_date > end_date:
            return JsonResponse({"error": "La fecha inicial no puede ser posterior a la final."}, status=400)
        if (end_date - start_date).days > 366:
            return JsonResponse(
                {"error": "El rango no puede superar 366 días. Descarga por tramos más cortos."},
                status=400,
            )

        # Fetch actions based on the scope
        if scope == 'team':
            actions = Action.objects.filter(team=team)
        elif scope == 'user':
            actions = Action.objects.filter(user=request.user)
        else:
            return JsonResponse({"error": "Invalid scope provided. Allowed scopes: 'team', 'user'."}, status=400)

        # El loop de mas abajo accede a lead, phone, medio, resultado y user de cada Action:
        # sin esto son 5 queries extra POR FILA (N+1) en vez de un solo join.
        actions = actions.select_related('lead', 'phone', 'medio', 'resultado', 'user')

        # Apply additional filters (by created_at date, and by cartera)
        if start_date:
            actions = actions.filter(created_at__gte=start_date)
        if end_date:
            # Limite EXCLUSIVO al dia siguiente: "created_at__lte=end_date" compara contra la
            # MEDIANOCHE de end_date (Django convierte el date a datetime 00:00), lo que dejaba
            # afuera cualquier gestion hecha durante el propio end_date (ej. "Hasta: hoy" no
            # traia nada de hoy). Asi se incluyen las 24 horas completas de end_date.
            actions = actions.filter(created_at__lt=end_date + datetime.timedelta(days=1))
        if cartera_id:
            actions = actions.filter(subcartera__cartera_id=cartera_id)

        logger.debug(f"{actions.count()} actions found for {scope}")

        # Handle case when no actions are found
        if not actions.exists():
            return JsonResponse({"error": "No actions found for the given criteria."}, status=404)

        # Generate Excel file for download
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = f'{scope.capitalize()} Actions'

        # Add headers to the spreadsheet
        headers = ['OP', 'RUT', 'ACCION', 'SUBESTADO', 'ESTADO', 'FECHA', 'COMENTARIO', 'TELEFONO', 'EMAIL',
                   'USER']
        sheet.append(headers)

        report_tz = datetime.timezone(datetime.timedelta(hours=-4))  # UTC-4, como pide el formato Galgo

        # Insert action data into the Excel sheet
        for action in actions:
            phone = str(action.phone) if action.phone else ''
            email = str(action.email) if action.email else ''
            rut = f"{action.lead.rut}{action.lead.dv}" if action.lead else ''

            sheet.append([
                action.lead.op if action.lead else '',  # OP
                rut,  # RUT (rut + dv concatenados)
                action.medio.nombre,  # ACCION
                action.get_target_display() or '',  # SUBESTADO
                action.resultado.nombre,  # ESTADO
                action.created_at.astimezone(report_tz).strftime("%Y-%m-%d %H:%M:%S"),  # FECHA (UTC-4)
                action.comment if action.comment else '',  # COMENTARIO
                phone,  # TELEFONO
                email,  # EMAIL
                action.user.username if action.user else ''  # USER
            ])

        # Save the workbook to a BytesIO object
        output = BytesIO()
        workbook.save(output)
        output.seek(0)

        # Create HTTP response for downloading the Excel file
        response = HttpResponse(
            content=output.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        filename_parts = [scope, 'actions']
        if cartera_id:
            cartera_obj = Cartera.objects.filter(pk=cartera_id).first()
            if cartera_obj:
                filename_parts.append(re.sub(r'\W+', '_', cartera_obj.nombre))
        response['Content-Disposition'] = f'attachment; filename={"_".join(filename_parts)}.xlsx'
        from configuracion.models import AccessLog, registrar_acceso
        registrar_acceso(request.user, AccessLog.EXPORTAR_GESTIONES, detail='Excel de gestiones')
        return response


SUPERVISOR_TYPES = ('admin', 'owner', 'supervisor')


class SupervisorRequiredMixin(LoginRequiredMixin):
    def dispatch(self, request, *args, **kwargs):
        userprofile = getattr(request.user, 'userprofile', None)
        if not userprofile or userprofile.user_type not in SUPERVISOR_TYPES:
            raise PermissionDenied
        return super().dispatch(request, *args, **kwargs)


class RecordingListView(SupervisorRequiredMixin, ListView):
    model = CallRecording
    template_name = 'actions/recordings_list.html'
    context_object_name = 'recordings'

    def get_limit(self):
        try:
            limit = int(self.request.GET.get('limit', 10))
        except ValueError:
            limit = 10
        return limit if limit in (10, 50, 100) else 10

    def get_queryset(self):
        # Alcance por cartera: un supervisor solo ve las grabaciones de sus carteras.
        qs = scope_por_lead(
            CallRecording.objects.select_related('lead__subcartera__cartera', 'user'), self.request.user
        ).order_by('-created_at')

        op = self.request.GET.get('op')
        if op:
            qs = qs.filter(lead__op__icontains=op)

        cartera_id = self.request.GET.get('cartera')
        if cartera_id:
            qs = qs.filter(lead__subcartera__cartera_id=cartera_id)

        fecha = self.request.GET.get('fecha')
        if fecha:
            parsed = parse_date(fecha)
            if parsed:
                qs = qs.filter(Q(call_date__date=parsed) | Q(call_date__isnull=True, created_at__date=parsed))

        return qs[:self.get_limit()]

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['limit'] = self.get_limit()
        context['carteras'] = carteras_visibles(self.request.user)
        context['selected_op'] = self.request.GET.get('op', '')
        context['selected_cartera'] = self.request.GET.get('cartera', '')
        context['selected_fecha'] = self.request.GET.get('fecha', '')
        context['es_admin'] = es_admin_owner(self.request.user)
        return context


class SyncRecordingsNowView(LoginRequiredMixin, View):
    """Botón de admin en Grabaciones para disparar la sincronización a mano (sin esperar el
    cron de cada 5 min). Despacha por Celery (.delay), NO inline: con muchos usuarios con
    llamadas pendientes, consultar/descargar de pbxip.cl uno por uno dentro del request puede
    superar el timeout de Azure (~230s) y dejar la pagina colgada. Al ir por Celery ademas
    respeta el rate_limit de sync_pbx_recordings_user (ver tasks.py), necesario cuando la cuenta
    MAESTRA (PBX_MASTER_EMAIL) es la que hace todas las consultas por todos los usuarios."""
    def post(self, request, *args, **kwargs):
        if not es_admin_owner(request.user):
            raise PermissionDenied
        from .tasks import sync_pbx_recordings_user, MIN_AGE_SECONDS, MAX_ATTEMPTS
        cutoff = timezone.now() - datetime.timedelta(seconds=MIN_AGE_SECONDS)
        user_ids = list(
            PendingPbxCall.objects.filter(
                resolved=False, requested_at__lte=cutoff, attempts__lt=MAX_ATTEMPTS
            ).order_by('user_id').values_list('user_id', flat=True).distinct()
        )
        for uid in user_ids:
            sync_pbx_recordings_user.delay(uid)
        if user_ids:
            messages.success(
                request,
                f'Sincronización en curso sobre {len(user_ids)} usuario(s) con llamadas pendientes. '
                'Las grabaciones nuevas aparecerán en el listado en los próximos minutos.'
            )
        else:
            messages.info(request, 'No había llamadas pendientes de sincronizar.')
        return redirect('actions:recordings_list')


class RecordingsExportTemplateView(SupervisorRequiredMixin, View):
    """Plantilla Excel para elegir qué grabaciones exportar en lote."""
    def get(self, request, *args, **kwargs):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Grabaciones'
        ws.append(['cartera', 'subcartera', 'op', 'fecha'])
        ws.append(['CARTERA-EJEMPLO', 'SUBCARTERA-EJEMPLO', 'OP-EJEMPLO', '2026-07-30'])

        output = BytesIO()
        wb.save(output)
        output.seek(0)
        response = HttpResponse(
            content=output.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        response['Content-Disposition'] = 'attachment; filename=plantilla_exportar_grabaciones.xlsx'
        return response


class RecordingsExportZipView(SupervisorRequiredMixin, View):
    """
    Recibe el Excel y encola el armado del ZIP en el WORKER de Celery (no lo arma en el proceso
    web): asi una descarga grande de grabaciones nunca frena el ingreso de gestiones ni se cae por
    timeout. El usuario queda en la lista de exports, donde ve el progreso y descarga cuando esta.
    """
    def post(self, request, *args, **kwargs):
        excel_file = request.FILES.get('excel_file')
        if not excel_file:
            messages.error(request, 'Debes subir un archivo Excel.')
            return redirect('actions:recordings_list')

        from .models import GrabacionesExportJob
        from .tasks import generar_zip_grabaciones
        job = GrabacionesExportJob.objects.create(
            solicitado_por=request.user,
            excel=excel_file,
        )
        # Se procesa en el worker; si Celery/Redis no responde, no se pierde el pedido (queda
        # PENDIENTE y la tarea periodica lo puede retomar / se reintenta manual).
        try:
            generar_zip_grabaciones.delay(job.pk)
        except Exception:
            logger.exception('No se pudo encolar generar_zip_grabaciones; queda pendiente.')

        messages.success(
            request,
            'Estamos generando el ZIP de grabaciones. Aparecerá para descargar en "Mis exports" '
            'en cuanto esté listo (no necesitas esperar en esta página).'
        )
        return redirect('actions:recordings_exports')


class RecordingsExportListView(SupervisorRequiredMixin, ListView):
    """Lista los exports de grabaciones del usuario (los suyos; admin/owner ven todos), con su
    estado y el link de descarga cuando estan listos."""
    template_name = 'actions/recordings_exports.html'
    context_object_name = 'jobs'
    paginate_by = 20

    def get_queryset(self):
        from .models import GrabacionesExportJob
        qs = GrabacionesExportJob.objects.select_related('solicitado_por')
        if not es_admin_owner(self.request.user):
            qs = qs.filter(solicitado_por=self.request.user)
        return qs


class RecordingsExportDownloadView(SupervisorRequiredMixin, View):
    """Descarga el ZIP ya generado de un job. Solo el dueño del job (o admin/owner) puede bajarlo."""
    def get(self, request, *args, **kwargs):
        from .models import GrabacionesExportJob
        job = get_object_or_404(GrabacionesExportJob, pk=kwargs.get('pk'))
        if job.solicitado_por_id != request.user.pk and not es_admin_owner(request.user):
            raise PermissionDenied
        if job.estado != GrabacionesExportJob.LISTO or not job.archivo:
            messages.error(request, 'Ese export todavía no está listo o no tiene archivo.')
            return redirect('actions:recordings_exports')

        from configuracion.models import AccessLog, registrar_acceso
        registrar_acceso(request.user, AccessLog.DESCARGAR_GRABACIONES, detail=f'ZIP de {job.total} grabación(es)')
        response = FileResponse(job.archivo.open('rb'), as_attachment=True,
                                filename=f'grabaciones_{job.pk}.zip', content_type='application/zip')
        return response


from .tanner_report import (
    TANNER_GESTOR_CODIGO, TANNER_ORIGEN_GESTION_DEFAULT, TANNER_TIPO_GESTION_DEFAULT,
    construir_lineas, gestiones_del_dia, nombre_archivo as _tanner_nombre_archivo,
    nombre_ejecutivo as _ejecutivo_nombre_tanner,
)


class TannerReportView(SupervisorRequiredMixin, View):
    """
    Genera el archivo de gestiones para Tanner tal como lo exige su instructivo: 14 columnas
    separadas por '|', sin encabezado, un solo dia por archivo. Se envia manualmente por correo a
    gestionescobranza@tanner.cl (no se automatiza el envio, solo se genera el archivo).

    Sin el parametro `subcartera`, es el reporte OFICIAL: TODO Tanner junto, sin importar quien lo
    baje -- el envio regulatorio a Tanner es UNO solo por dia para toda la cartera, asi que
    cualquier supervisor (no solo admin/owner) puede generar el archivo completo. Con `subcartera`
    (opcional, un id), un supervisor puede bajar SOLO esa subcartera suya.

    El formato vive en actions/tanner_report.py, compartido con el reporte por RANGO de fechas
    (que corre en el worker): es un archivo regulatorio, no puede haber dos implementaciones.
    """

    def get(self, request, *args, **kwargs):
        fecha_str = request.GET.get('fecha')
        if not fecha_str:
            return JsonResponse({'error': 'Debes indicar una fecha (YYYY-MM-DD).'}, status=400)

        fecha = parse_date(fecha_str)
        if not fecha:
            return JsonResponse({'error': 'Fecha invalida. Usa el formato YYYY-MM-DD.'}, status=400)

        subcartera_id = request.GET.get('subcartera')
        actions = gestiones_del_dia(fecha, request.user, subcartera_id)
        if not actions.exists():
            return JsonResponse({'error': 'No hay gestiones de Tanner para esa fecha.'}, status=404)

        content = '\r\n'.join(construir_lineas(actions)) + '\r\n'
        subcartera = Subcartera.objects.filter(pk=subcartera_id).first() if subcartera_id else None
        filename = _tanner_nombre_archivo(fecha, subcartera)

        response = HttpResponse(content, content_type='text/plain; charset=utf-8')
        response['Content-Disposition'] = f'attachment; filename={filename}'
        from configuracion.models import AccessLog, registrar_acceso
        registrar_acceso(request.user, AccessLog.EXPORTAR_TANNER, detail=f"Fecha {fecha:%d-%m-%Y}")
        return response


class TannerReportRangeView(SupervisorRequiredMixin, View):
    """
    Pide el reporte Tanner de un RANGO de fechas: el worker arma un ZIP con un .txt por dia (con
    su nombre oficial, listo para enviar) mas un consolidado. Encolar en vez de generar en el
    request: un rango largo supera el limite de espera del proxy de Azure.
    """

    def post(self, request, *args, **kwargs):
        from .models import ReporteTannerJob
        from .tasks import generar_reporte_tanner_rango

        desde = parse_date(request.POST.get('desde', ''))
        hasta = parse_date(request.POST.get('hasta', ''))
        if not desde or not hasta:
            messages.error(request, 'Indica la fecha de inicio y la de termino.')
            return redirect('actions:index')
        if desde > hasta:
            messages.error(request, 'La fecha de inicio no puede ser posterior a la de termino.')
            return redirect('actions:index')

        subcartera = None
        subcartera_id = request.POST.get('subcartera')
        if subcartera_id:
            # Mismo criterio que la descarga de un dia: solo una subcartera que el usuario
            # realmente supervise (admin/owner ven todas).
            subcartera = subcarteras_visibles(request.user).filter(pk=subcartera_id).first()
            if not subcartera:
                messages.error(request, 'Esa subcartera no existe o no la supervisas.')
                return redirect('actions:index')

        job = ReporteTannerJob.objects.create(
            solicitado_por=request.user, fecha_desde=desde, fecha_hasta=hasta, subcartera=subcartera,
        )
        try:
            generar_reporte_tanner_rango.delay(job.pk)
        except Exception:
            logger.exception('No se pudo encolar generar_reporte_tanner_rango; queda pendiente.')

        messages.success(
            request,
            'Estamos generando el reporte Tanner del rango. Aparecera para descargar en '
            '"Reportes Tanner por rango" en cuanto este listo (no necesitas esperar aqui).'
        )
        return redirect('actions:tanner_rango_list')


class TannerReportRangeListView(SupervisorRequiredMixin, ListView):
    """Lista los reportes Tanner por rango pedidos (los propios; admin/owner ven todos)."""
    template_name = 'actions/tanner_rango_list.html'
    context_object_name = 'jobs'
    paginate_by = 20

    def get_queryset(self):
        from .models import ReporteTannerJob
        qs = ReporteTannerJob.objects.select_related('solicitado_por', 'subcartera')
        if not es_admin_owner(self.request.user):
            qs = qs.filter(solicitado_por=self.request.user)
        return qs


class TannerReportRangeDownloadView(SupervisorRequiredMixin, View):
    """Descarga el ZIP ya generado. Solo el dueno del job (o admin/owner)."""

    def get(self, request, *args, **kwargs):
        from .models import ReporteTannerJob
        job = get_object_or_404(ReporteTannerJob, pk=kwargs.get('pk'))
        if job.solicitado_por_id != request.user.pk and not es_admin_owner(request.user):
            raise PermissionDenied
        if job.estado != ReporteTannerJob.LISTO or not job.archivo:
            messages.error(request, 'Ese reporte todavia no esta listo.')
            return redirect('actions:tanner_rango_list')

        from configuracion.models import AccessLog, registrar_acceso
        registrar_acceso(
            request.user, AccessLog.EXPORTAR_TANNER,
            detail=f"Rango {job.fecha_desde:%d-%m-%Y} a {job.fecha_hasta:%d-%m-%Y}",
        )
        return FileResponse(
            job.archivo.open('rb'), as_attachment=True,
            filename=f'tanner_{job.fecha_desde:%Y%m%d}_a_{job.fecha_hasta:%Y%m%d}.zip',
            content_type='application/zip',
        )


def _rango_semana_mes(today):
    """Devuelve (inicio_semana, fin_semana, inicio_mes, inicio_mes_siguiente)."""
    week_start = today - datetime.timedelta(days=today.weekday())
    week_end = week_start + datetime.timedelta(days=6)
    month_start = today.replace(day=1)
    if month_start.month == 12:
        next_month = month_start.replace(year=month_start.year + 1, month=1)
    else:
        next_month = month_start.replace(month=month_start.month + 1)
    return week_start, week_end, month_start, next_month


def _resumen(qs, campo_monto='monto'):
    agg = qs.aggregate(n=Count('id'), total=Sum(campo_monto))
    return {'count': agg['n'] or 0, 'total': agg['total'] or 0}


class PaymentCommitmentListView(LoginRequiredMixin, View):
    template_name = 'actions/commitments_list.html'

    def get(self, request, *args, **kwargs):
        selected_op = request.GET.get('op', '').strip()
        selected_cartera = request.GET.get('cartera', '').strip()
        selected_fecha = request.GET.get('fecha', '').strip()
        es_super = _es_supervisor(request.user)

        # Alcance por rol: cobrador -> sus clientes; supervisor -> sus carteras; admin/owner -> todo.
        # Solo vigentes: uno editado o roto sale de tarjetas/tabla (sigue existiendo para
        # auditoria, ver actions/commitment_lifecycle.py).
        base_all = scope_por_lead(PaymentCommitment.objects.filter(vigente=True), request.user)

        # Filtros que NO son de cartera (op/fecha) — se aplican tanto a las tarjetas como a la tabla.
        base_filtered = base_all.select_related('lead', 'subcartera__cartera', 'created_by')
        if selected_op:
            base_filtered = base_filtered.filter(lead__op__icontains=selected_op)
        if selected_fecha:
            parsed = parse_date(selected_fecha)
            if parsed:
                base_filtered = base_filtered.filter(fecha_compromiso=parsed)

        # Tarjetas: una por cartera (solo las que tienen compromisos dentro del alcance del usuario).
        carteras_launcher = []
        cartera_ids = base_all.values_list('subcartera__cartera_id', flat=True).distinct()
        for cartera in Cartera.objects.filter(id__in=cartera_ids):
            resumen = _resumen(base_filtered.filter(subcartera__cartera=cartera))
            carteras_launcher.append({
                'id': cartera.id, 'nombre': cartera.nombre,
                'count': resumen['count'], 'total': resumen['total'],
                'active': selected_cartera == str(cartera.id),
            })
        gran_total = _resumen(base_filtered)

        # Tabla: los mismos filtros, más el de cartera si se eligió una tarjeta.
        tabla_qs = base_filtered
        if selected_cartera:
            tabla_qs = tabla_qs.filter(subcartera__cartera_id=selected_cartera)
        tabla_qs = tabla_qs.order_by('-fecha_compromiso', '-created_at')

        # "Mostrar: 10/50/100/Todos" -- mismo patron que la lista de clientes (LeadListView).
        limit_raw = request.GET.get('limit', '10')
        if limit_raw == 'todos':
            limit = None
        else:
            try:
                limit = int(limit_raw)
            except ValueError:
                limit = 10
            if limit not in (10, 50, 100):
                limit = 10
        tabla_compromisos = list(tabla_qs if limit is None else tabla_qs[:limit])

        def _url_limit(value):
            params = request.GET.copy()
            params['limit'] = value
            return '?' + params.urlencode()

        # Resumen por fecha de compromiso (hoy / esta semana / este mes), dentro del alcance.
        today = timezone.localdate()
        week_start, week_end, month_start, next_month = _rango_semana_mes(today)
        context = {
            'carteras_launcher': carteras_launcher,
            'gran_total': gran_total,
            'tabla_compromisos': tabla_compromisos,
            'resumen_hoy': _resumen(base_all.filter(fecha_compromiso=today)),
            'resumen_semana': _resumen(base_all.filter(fecha_compromiso__gte=week_start, fecha_compromiso__lte=week_end)),
            'resumen_mes': _resumen(base_all.filter(fecha_compromiso__gte=month_start, fecha_compromiso__lt=next_month)),
            'selected_op': selected_op,
            'selected_cartera': selected_cartera,
            'selected_fecha': selected_fecha,
            'limit': 'todos' if limit is None else limit,
            'url_limit_10': _url_limit(10),
            'url_limit_50': _url_limit(50),
            'url_limit_100': _url_limit(100),
            'url_limit_todos': _url_limit('todos'),
            'es_supervisor': es_super,
        }
        return render(request, self.template_name, context)

    def post(self, request, *args, **kwargs):
        from .commitment_lifecycle import editar, marcar_roto

        accion = request.POST.get('accion')
        next_url = request.POST.get('next') or 'actions:commitments_list'
        commitment = get_object_or_404(
            scope_por_lead(
                PaymentCommitment.objects.select_related('action', 'lead'), request.user
            ),
            pk=request.POST.get('commitment_id'),
        )
        if not commitment.vigente:
            messages.error(request, 'Ese compromiso ya fue retirado; recarga la página.')
            return redirect(next_url)

        if accion == 'editar':
            nueva_fecha = parse_date(request.POST.get('nueva_fecha', ''))
            monto_raw = request.POST.get('nuevo_monto', '').strip()
            if not nueva_fecha:
                messages.error(request, 'Ingresa una fecha de compromiso válida.')
                return redirect(next_url)
            # localdate(), no date.today(): ver nota en actions/forms.py (ActionForm.clean).
            if nueva_fecha < timezone.localdate():
                messages.error(request, 'La fecha de compromiso no puede ser anterior a hoy.')
                return redirect(next_url)
            try:
                nuevo_monto = int(monto_raw) if monto_raw else None
            except ValueError:
                messages.error(request, 'El monto debe ser un número.')
                return redirect(next_url)
            editar(
                commitment, nueva_fecha, nuevo_monto, request.user,
                comentario=request.POST.get('comentario', '').strip(),
            )
            messages.success(
                request,
                f'Compromiso de {commitment.lead.op} editado: nueva fecha {nueva_fecha:%d-%m-%Y}.'
            )

        elif accion == 'roto':
            marcar_roto(commitment, request.user, comentario=request.POST.get('comentario', '').strip())
            messages.success(
                request,
                f'Compromiso de {commitment.lead.op} marcado como roto. El cliente queda en estado Contactado.'
            )

        else:
            messages.error(request, 'Acción inválida.')

        return redirect(next_url)


class BrokenCommitmentsListView(LoginRequiredMixin, View):
    """
    Compromisos de pago marcados como rotos (ver actions/commitment_lifecycle.marcar_roto): no
    generan gestion, son puro cambio de status a Contactado -- esta pantalla es su propio
    historial, separado de "Compromisos de pago" (que solo muestra vigentes). Se purgan solos
    despues de RetentionSettings.dias_retencion_compromisos_rotos dias
    (actions.tasks.purgar_compromisos_rotos), asi que esta lista queda acotada a esa ventana sola.
    """
    template_name = 'actions/commitments_broken_list.html'

    def get(self, request, *args, **kwargs):
        from suspensiones.models import RetentionSettings

        selected_op = request.GET.get('op', '').strip()
        selected_cartera = request.GET.get('cartera', '').strip()

        dias_retencion = RetentionSettings.get_solo().dias_retencion_compromisos_rotos
        corte = timezone.now() - datetime.timedelta(days=dias_retencion)

        # Alcance por rol: cobrador -> sus clientes; supervisor -> sus carteras; admin/owner -> todo.
        base_all = scope_por_lead(
            PaymentCommitment.objects.filter(
                motivo_retiro=PaymentCommitment.MOTIVO_ROTO, retirado_at__gte=corte,
            ),
            request.user,
        ).select_related('lead', 'subcartera__cartera', 'retirado_por')

        cartera_ids = base_all.values_list('subcartera__cartera_id', flat=True).distinct()
        carteras_choices = Cartera.objects.filter(id__in=cartera_ids)

        qs = base_all
        if selected_op:
            qs = qs.filter(lead__op__icontains=selected_op)
        if selected_cartera:
            qs = qs.filter(subcartera__cartera_id=selected_cartera)
        qs = qs.order_by('-retirado_at')
        total_filtrado = qs.count()

        limit_raw = request.GET.get('limit', '10')
        if limit_raw == 'todos':
            limit = None
        else:
            try:
                limit = int(limit_raw)
            except ValueError:
                limit = 10
            if limit not in (10, 50, 100):
                limit = 10
        tabla = list(qs if limit is None else qs[:limit])

        def _url_limit(value):
            params = request.GET.copy()
            params['limit'] = value
            return '?' + params.urlencode()

        context = {
            'tabla': tabla,
            'total': total_filtrado,
            'dias_retencion': dias_retencion,
            'selected_op': selected_op,
            'selected_cartera': selected_cartera,
            'carteras_choices': carteras_choices,
            'limit': 'todos' if limit is None else limit,
            'url_limit_10': _url_limit(10),
            'url_limit_50': _url_limit(50),
            'url_limit_100': _url_limit(100),
            'url_limit_todos': _url_limit('todos'),
        }
        return render(request, self.template_name, context)


class PaymentCommitmentExportExcelView(SupervisorRequiredMixin, View):
    def get(self, request, *args, **kwargs):
        qs = PaymentCommitment.objects.filter(vigente=True).select_related(
            'lead', 'subcartera__cartera', 'created_by'
        ).order_by('-fecha_compromiso', '-created_at')

        op = request.GET.get('op')
        if op:
            qs = qs.filter(lead__op__icontains=op)

        cartera_id = request.GET.get('cartera')
        if cartera_id:
            qs = qs.filter(subcartera__cartera_id=cartera_id)

        fecha = request.GET.get('fecha')
        if fecha:
            parsed = parse_date(fecha)
            if parsed:
                qs = qs.filter(fecha_compromiso=parsed)

        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = 'Compromisos de pago'
        sheet.append(['Cartera', 'Subcartera', 'ID', 'Cliente', 'Fecha compromiso', 'Monto', 'Comentario', 'Ejecutivo'])

        for c in qs:
            sheet.append([
                c.subcartera.cartera.nombre,
                c.subcartera.nombre,
                c.lead.op,
                c.lead.name,
                c.fecha_compromiso.strftime('%d-%m-%Y') if c.fecha_compromiso else '',
                float(c.monto) if c.monto is not None else '',
                c.comentario or '',
                c.created_by.username if c.created_by else '',
            ])

        output = BytesIO()
        workbook.save(output)
        output.seek(0)

        response = HttpResponse(
            content=output.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        response['Content-Disposition'] = 'attachment; filename=compromisos_de_pago.xlsx'
        from configuracion.models import AccessLog, registrar_acceso
        registrar_acceso(request.user, AccessLog.EXPORTAR_COMPROMISOS, detail='Excel de compromisos')
        return response


# ------------------ Carga masiva de gestiones (campanas: IVR, SMS, discador, etc.) ------------------

# La validacion y el guardado viven en actions/bulk_gestiones.py y corren en el WORKER de Celery
# (actions.tasks.procesar_carga_gestiones): un Excel de decenas de miles de filas tarda mas que el
# limite de espera del proxy de Azure, que devolvia 504 al navegador MIENTRAS el servidor seguia
# guardando -- el usuario veia "error", reintentaba, y las gestiones quedaban duplicadas.
from .bulk_gestiones import BULK_TEMPLATE_EXAMPLE, BULK_TEMPLATE_HEADERS


class BulkActionUploadView(SupervisorRequiredMixin, View):
    """Recibe el Excel de gestiones y encola su procesamiento; no guarda nada en el request."""
    template_name = 'actions/bulk_upload.html'

    def get(self, request, *args, **kwargs):
        from .models import CargaGestionesJob
        jobs = CargaGestionesJob.objects.select_related('solicitado_por')
        if not es_admin_owner(request.user):
            jobs = jobs.filter(solicitado_por=request.user)
        return render(request, self.template_name, {'jobs': jobs[:20]})

    def post(self, request, *args, **kwargs):
        from .models import CargaGestionesJob
        from .tasks import procesar_carga_gestiones

        excel_file = request.FILES.get('excel_file')
        if not excel_file:
            messages.error(request, 'Debes subir un archivo Excel.')
            return redirect('actions:bulk_upload')

        job = CargaGestionesJob.objects.create(
            solicitado_por=request.user, excel=excel_file,
        )
        # Si Celery/Redis no responde, el pedido NO se pierde: queda PENDIENTE en la lista.
        try:
            procesar_carga_gestiones.delay(job.pk)
        except Exception:
            logger.exception('No se pudo encolar procesar_carga_gestiones; queda pendiente.')

        messages.success(
            request,
            'Recibimos el archivo y lo estamos procesando. El resultado aparece abajo en '
            '"Cargas recientes" en cuanto termine (no necesitas esperar en esta pagina). '
            'Si alguna fila tiene error, no se carga ninguna y podras descargar el Excel con el detalle.'
        )
        return redirect('actions:bulk_upload')


class BulkActionErroresDownloadView(SupervisorRequiredMixin, View):
    """Descarga el Excel de errores de una carga de gestiones rechazada."""

    def get(self, request, *args, **kwargs):
        from .models import CargaGestionesJob
        job = get_object_or_404(CargaGestionesJob, pk=kwargs.get('pk'))
        if job.solicitado_por_id != request.user.pk and not es_admin_owner(request.user):
            raise PermissionDenied
        if not job.archivo_errores:
            messages.error(request, 'Esa carga no tiene un archivo de errores.')
            return redirect('actions:bulk_upload')
        return FileResponse(
            job.archivo_errores.open('rb'), as_attachment=True,
            filename=f'errores_gestiones_{job.pk}.xlsx',
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )


class BulkActionTemplateView(SupervisorRequiredMixin, View):
    """Descarga la plantilla Excel para la carga masiva de gestiones."""

    def get(self, request, *args, **kwargs):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Gestiones'
        ws.append(BULK_TEMPLATE_HEADERS)
        ws.append(BULK_TEMPLATE_EXAMPLE)

        output = BytesIO()
        wb.save(output)
        output.seek(0)
        response = HttpResponse(
            content=output.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        response['Content-Disposition'] = 'attachment; filename=plantilla_carga_gestiones.xlsx'
        return response


# ------------------ Pagos (independiente de las gestiones) ------------------

def _es_supervisor(user):
    profile = getattr(user, 'userprofile', None)
    return bool(profile and profile.user_type in ('admin', 'owner', 'supervisor'))


class PaymentCreateView(LoginRequiredMixin, View):
    """Formulario de pago en popup (se abre en el modal, igual que gestiones/notas)."""
    template_name = 'actions/payment_form.html'

    def get(self, request, lead_id, *args, **kwargs):
        # Sin esto, cualquier usuario autenticado podia registrar un pago sobre CUALQUIER lead
        # pasando su id -- fuera de alcance -> 404, mismo criterio que el resto de la app.
        lead = get_object_or_404(leads_visibles(request.user), id=lead_id)
        form = PaymentForm(initial={'fecha': timezone.localdate()})
        return render(request, self.template_name, {'lead': lead, 'form': form})

    def post(self, request, lead_id, *args, **kwargs):
        lead = get_object_or_404(leads_visibles(request.user), id=lead_id)
        form = PaymentForm(request.POST, request.FILES)
        if form.is_valid():
            pago = form.save(commit=False)
            pago.lead = lead
            pago.created_by = request.user
            pago.save()  # subcartera se hereda del lead en Payment.save()
            messages.success(request, 'Pago registrado correctamente.')
            return redirect('actions:popup_close')
        return render(request, self.template_name, {'lead': lead, 'form': form})


class PaymentListView(LoginRequiredMixin, View):
    """Dashboard de pagos: buscador para ingresar un pago, tarjetas por cartera y tabla filtrable."""
    template_name = 'actions/payments_list.html'

    def get(self, request, *args, **kwargs):
        es_super = _es_supervisor(request.user)
        selected_cartera = request.GET.get('cartera', '').strip()

        # Alcance por rol: cobrador -> sus clientes; supervisor -> sus carteras; admin/owner -> todo.
        payments = scope_por_lead(
            Payment.objects.select_related('lead', 'subcartera__cartera', 'created_by'), request.user
        ).order_by('-fecha', '-created_at')

        # Buscador de cliente para registrar un pago.
        q = request.GET.get('q', '').strip()
        lead_results = []
        if q:
            leads = leads_visibles(request.user, base=Lead.objects.select_related('subcartera__cartera'))
            lead_results = list(
                leads.filter(Q(op__icontains=q) | Q(rut__icontains=q) | Q(name__icontains=q))[:20]
            )

        # Sumatoria de pagos de ESTE MES: total y por cartera (N° y monto).
        today = timezone.localdate()
        _, _, month_start, next_month = _rango_semana_mes(today)
        mes_qs = payments.filter(fecha__gte=month_start, fecha__lt=next_month)
        resumen_mes_total = _resumen(mes_qs, 'monto')

        # Tarjetas: una por cartera (sobre los pagos de este mes), sirven de filtro para la tabla.
        carteras_launcher = []
        for cartera in Cartera.objects.all():
            resumen = _resumen(mes_qs.filter(subcartera__cartera=cartera), 'monto')
            carteras_launcher.append({
                'id': cartera.id, 'nombre': cartera.nombre,
                'count': resumen['count'], 'total': resumen['total'],
                'active': selected_cartera == str(cartera.id),
            })

        # Pagos historicos: busqueda libre por OP y/o fecha exacta sobre TODOS los pagos (no solo
        # el mes) -- antes la unica forma de encontrar un pago viejo era entrar al lead uno por
        # uno. Independiente del filtro de cartera de arriba (se combinan).
        hist_op = request.GET.get('hist_op', '').strip()
        hist_fecha = request.GET.get('hist_fecha', '').strip()
        tabla_qs = payments
        if selected_cartera:
            tabla_qs = tabla_qs.filter(subcartera__cartera_id=selected_cartera)
        if hist_op:
            tabla_qs = tabla_qs.filter(lead__op__icontains=hist_op)
        if hist_fecha:
            parsed_fecha = parse_date(hist_fecha)
            if parsed_fecha:
                tabla_qs = tabla_qs.filter(fecha=parsed_fecha)

        # "Mostrar: 10/50/100/Todos" -- mismo patron que Compromisos de pago.
        limit_raw = request.GET.get('limit', '10')
        if limit_raw == 'todos':
            limit = None
        else:
            try:
                limit = int(limit_raw)
            except ValueError:
                limit = 10
            if limit not in (10, 50, 100):
                limit = 10
        tabla_pagos = list(tabla_qs if limit is None else tabla_qs[:limit])

        def _url_limit(value):
            params = request.GET.copy()
            params['limit'] = value
            return '?' + params.urlencode()

        context = {
            'q': q,
            'lead_results': lead_results,
            'carteras_launcher': carteras_launcher,
            'resumen_mes_total': resumen_mes_total,
            'tabla_pagos': tabla_pagos,
            'selected_cartera': selected_cartera,
            'hist_op': hist_op,
            'hist_fecha': hist_fecha,
            'limit': 'todos' if limit is None else limit,
            'url_limit_10': _url_limit(10),
            'url_limit_50': _url_limit(50),
            'url_limit_100': _url_limit(100),
            'url_limit_todos': _url_limit('todos'),
        }
        return render(request, self.template_name, context)


NC_EXTERNO = 'ZONA SUR'  # columna "Externo" fija en el reporte de Nuevo Capital


def _ejecutivo_nombre(user):
    if not user:
        return ''
    full = user.get_full_name().strip()
    return (full or user.username).upper()


class NuevoCapitalReportView(SupervisorRequiredMixin, View):
    """
    Genera el reporte de gestiones de Nuevo Capital (.xlsx, 16 columnas) para un dia.
    Formato de las columnas segun su instructivo (Reporte Salida NC):
    Operacion, RUT, Dv, FechaGest, HoraGest, Usuario, Accion, Sub Estado, Estado,
    Comentario, Fecha Compromiso, Monto Compromiso, Telefono, eMail, Origen Gestion(In=1/Out=2),
    Externo(ZONA SUR).

    Como en Tanner, es el reporte OFICIAL: la cartera COMPLETA para el dia, sin acotar por
    scope_por_lead -- el envio regulatorio es uno solo por dia, no por subcartera, asi que
    cualquier supervisor (no solo admin/owner) puede generarlo.
    """

    HEADERS = [
        'Operación', 'RUT', 'Dv', 'FechaGest', 'Hora Gest', 'Usuario', 'Accion',
        'Sub Estado', 'Estado', 'Comentario', 'Fecha Compromiso', 'Monto Compromiso',
        'Telefono', 'eMail', 'Origen Gestión', 'Externo',
    ]

    def get(self, request, *args, **kwargs):
        fecha_str = request.GET.get('fecha')
        if not fecha_str:
            return JsonResponse({'error': 'Debes indicar una fecha (YYYY-MM-DD).'}, status=400)
        fecha = parse_date(fecha_str)
        if not fecha:
            return JsonResponse({'error': 'Fecha inválida. Usa el formato YYYY-MM-DD.'}, status=400)

        report_tz = datetime.timezone(datetime.timedelta(hours=-4))
        start = datetime.datetime.combine(fecha, datetime.time.min, tzinfo=report_tz)
        end = datetime.datetime.combine(fecha, datetime.time.max, tzinfo=report_tz)

        actions = Action.objects.filter(
            subcartera__cartera__nombre__iexact='Nuevo Capital',
            created_at__gte=start,
            created_at__lte=end,
        ).select_related('lead', 'medio', 'resultado', 'user', 'phone')

        if not actions.exists():
            return JsonResponse({'error': 'No hay gestiones de Nuevo Capital para esa fecha.'}, status=404)

        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = 'Reporte Salida NC'
        sheet.append(self.HEADERS)

        for action in actions:
            lead = action.lead
            local_dt = action.created_at.astimezone(report_tz)
            origen = '1' if action.medio.es_inbound else '2'  # In=1 / Out=2
            sub_estado = action.resultado.tipo_contacto or ''
            compromiso = action.fecha_compromiso.strftime('%d-%m-%Y') if action.fecha_compromiso else ''

            sheet.append([
                lead.op,
                lead.rut,
                lead.dv,
                local_dt.strftime('%d-%m-%Y'),
                local_dt.strftime('%H:%M:%S'),
                _ejecutivo_nombre(action.user),
                action.medio.nombre,
                sub_estado,
                action.resultado.nombre,
                action.comment or '',
                compromiso,
                action.monto_compromiso if action.monto_compromiso else '',
                _format_tanner_phone(action.phone.phone_number) if action.phone else '',
                action.email or '',
                origen,
                NC_EXTERNO,
            ])

        output = BytesIO()
        workbook.save(output)
        output.seek(0)

        filename = f"Reporte_Gestiones_NuevoCapital_{fecha.strftime('%Y%m%d')}.xlsx"
        response = HttpResponse(
            content=output.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        response['Content-Disposition'] = f'attachment; filename={filename}'
        from configuracion.models import AccessLog, registrar_acceso
        registrar_acceso(request.user, AccessLog.EXPORTAR_NUEVOCAPITAL, detail=f"Fecha {fecha:%d-%m-%Y}")
        return response