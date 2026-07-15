from django.core.management.base import BaseCommand, CommandError
from openpyxl import load_workbook

from actions.models import Medio, Resultado
from cartera.models import Cartera

CANAL_POR_MEDIO = {
    'WHATSAPP': Medio.CANAL_TELEFONO,
    'TELEFONICO': Medio.CANAL_TELEFONO,
    'SMS': Medio.CANAL_TELEFONO,
    'EMAIL': Medio.CANAL_EMAIL,
}

# Medios que representan una llamada de voz real por la central (unicos con grabacion en pbxip.cl).
MEDIOS_LLAMADA = {'TELEFONICO', 'LLAMADA', 'LLAMADA TELEFONICA', 'VOZ'}


class Command(BaseCommand):
    help = (
        "Importa el arbol de gestiones (medios y resultados) de una cartera desde un Excel "
        "con columnas: Medio de contacto, Estado del contacto, Contactabilidad, DEFAULT, CREA COMPROMISO. "
        "Los datos empiezan en la fila 3 (las 2 primeras son encabezados). Medio y Resultado se "
        "guardan como catalogos independientes de la cartera (un mismo resultado puede repetirse "
        "bajo varios medios en el Excel de origen -- se deduplica por nombre)."
    )

    def add_arguments(self, parser):
        parser.add_argument('cartera_nombre')
        parser.add_argument('excel_path')

    def handle(self, *args, **options):
        cartera_nombre = options['cartera_nombre']
        excel_path = options['excel_path']

        try:
            cartera = Cartera.objects.get(nombre__iexact=cartera_nombre)
        except Cartera.DoesNotExist:
            raise CommandError(f"No existe la cartera '{cartera_nombre}'. Creala primero en /dashboard/carteras/.")

        wb = load_workbook(excel_path, data_only=True)
        ws = wb.active

        medios_creados = 0
        filas_omitidas = 0
        resultados_agg = {}

        for row in ws.iter_rows(min_row=3, values_only=True):
            if not row or len(row) < 2:
                continue
            medio_nombre, resultado_nombre, contactabilidad, es_default, crea_compromiso = (list(row) + [None] * 5)[:5]

            if not medio_nombre or not resultado_nombre:
                filas_omitidas += 1
                continue

            medio_nombre = str(medio_nombre).strip()
            resultado_nombre = str(resultado_nombre).strip()
            canal = CANAL_POR_MEDIO.get(medio_nombre.upper(), Medio.CANAL_TELEFONO)

            medio, medio_created = Medio.objects.get_or_create(
                cartera=cartera, nombre=medio_nombre,
                defaults={'canal': canal, 'es_llamada': medio_nombre.upper() in MEDIOS_LLAMADA}
            )
            if medio_created:
                medios_creados += 1

            contactabilidad_norm = (
                Resultado.CON_CONTACTO
                if 'CON CONTACTO' in str(contactabilidad or '').upper()
                else Resultado.SIN_CONTACTO
            )
            crea_compromiso_bool = bool(crea_compromiso)

            # Un mismo resultado (por nombre) puede aparecer bajo varios medios en el Excel de
            # origen (ej. "COMPROMISO DE PAGO" bajo WHATSAPP, EMAIL y TELEFONICO). Como el
            # Resultado ahora pertenece a la cartera, no al medio, se agregan (merge) todas las
            # apariciones antes de crear el catalogo final.
            agg = resultados_agg.setdefault(resultado_nombre, {
                'contactabilidad': contactabilidad_norm,
                'es_default': False,
                'crea_compromiso': False,
                'es_llamada': False,
            })
            agg['es_default'] = agg['es_default'] or bool(es_default)
            agg['crea_compromiso'] = agg['crea_compromiso'] or crea_compromiso_bool
            agg['es_llamada'] = agg['es_llamada'] or medio.es_llamada
            if contactabilidad_norm == Resultado.CON_CONTACTO:
                agg['contactabilidad'] = Resultado.CON_CONTACTO

        resultados_creados, resultados_actualizados = 0, 0
        for nombre, agg in resultados_agg.items():
            requiere_fecha_pago = agg['crea_compromiso'] or 'PAGO' in nombre.upper()

            resultado, resultado_created = Resultado.objects.get_or_create(
                cartera=cartera, nombre=nombre,
            )
            resultado.contactabilidad = agg['contactabilidad']
            resultado.es_default = agg['es_default']
            resultado.crea_compromiso = agg['crea_compromiso']
            resultado.requiere_fecha_pago = requiere_fecha_pago
            if resultado_created:
                # Solo se fija un valor inicial al crearlo: un supervisor puede curar esto
                # manualmente despues (que resultados de llamada guardan grabacion), y no
                # queremos que una re-importacion del arbol le pise su decision.
                resultado.descarga_grabacion = (
                    agg['es_llamada'] and agg['contactabilidad'] == Resultado.CON_CONTACTO
                )
                resultados_creados += 1
            else:
                resultados_actualizados += 1
            resultado.save()

        self.stdout.write(self.style.SUCCESS(
            f"Cartera '{cartera.nombre}': {medios_creados} medio(s) nuevo(s), "
            f"{resultados_creados} resultado(s) nuevo(s), {resultados_actualizados} actualizado(s), "
            f"{filas_omitidas} fila(s) omitida(s)."
        ))
