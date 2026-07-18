"""
Carga el arbol de gestiones de Nuevo Capital parseando su "Paleta Respuestas" (.xlsx).

A diferencia de Tanner (que trae tablas de codigos en un instructivo Word), Nuevo Capital
entrega un Excel con una fila por combinacion valida de:
  Accion (col C) | Sub Estado (col D) | Estado (col E) | In/Out Bound (col G)

Mapeo a nuestro modelo:
  - Accion  -> Medio (con es_inbound derivado de la col "In Out Bound" y del sufijo RECIBIDO)
  - Estado  -> Resultado.nombre
  - Sub Estado -> Resultado.tipo_contacto (DIRECTO / DIRECTO AVAL / INDIRECTO / SIN CONTACTO /
                  ACCION MASIVA)

Nuevo Capital no usa codigos numericos ni para medio ni para resultado (su reporte va con
nombres), asi que 'codigo' queda vacio.
"""
import re

from django.core.management.base import BaseCommand, CommandError
from openpyxl import load_workbook

from actions.models import Medio, Resultado
from cartera.models import Cartera

# Canal y si es llamada de voz, por accion (nombre normalizado en mayusculas).
CANAL_EMAIL_ACCIONES = {'CORREO', 'CORREO RECIBIDO'}
LLAMADA_ACCIONES = {'LLAMADA', 'LLAMADA RECIBIDA', 'IVR', 'IVR AUDIO'}

# tipo_contacto (Sub Estado) que implica contacto efectivo con una persona.
TIPOS_CON_CONTACTO = {'DIRECTO', 'DIRECTO AVAL', 'INDIRECTO'}


class Command(BaseCommand):
    help = "Carga el arbol de gestiones de Nuevo Capital desde su Paleta Respuestas (.xlsx)."

    def add_arguments(self, parser):
        parser.add_argument('excel_path')

    def handle(self, *args, **options):
        try:
            cartera = Cartera.objects.get(nombre__iexact='Nuevo Capital')
        except Cartera.DoesNotExist:
            raise CommandError("No existe la cartera 'Nuevo Capital'. Creala primero en /dashboard/carteras/.")

        wb = load_workbook(options['excel_path'], data_only=True)
        ws = wb.active

        # Primero recolectamos todo para poder deduplicar medios y resultados.
        medios_info = {}      # nombre -> {'canal', 'es_llamada', 'es_inbound'}
        resultados_info = {}  # (nombre, tipo_contacto) -> {'con_contacto', 'es_llamada'}

        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row or len(row) < 7:
                continue
            accion, sub_estado, estado, in_out = row[2], row[3], row[4], row[6]
            if not accion or not estado:
                continue

            accion = str(accion).strip()
            sub_estado = str(sub_estado or '').strip()
            estado = str(estado).strip()
            accion_upper = accion.upper()

            es_inbound = 'RECIBID' in accion_upper or 'in' in str(in_out or '').strip().lower()
            canal = Medio.CANAL_EMAIL if accion_upper in CANAL_EMAIL_ACCIONES else Medio.CANAL_TELEFONO
            es_llamada = accion_upper in LLAMADA_ACCIONES

            m = medios_info.setdefault(accion, {'canal': canal, 'es_llamada': es_llamada, 'es_inbound': es_inbound})
            m['es_inbound'] = m['es_inbound'] or es_inbound
            m['es_llamada'] = m['es_llamada'] or es_llamada

            con_contacto = sub_estado.upper() in TIPOS_CON_CONTACTO
            r = resultados_info.setdefault((estado, sub_estado), {'con_contacto': con_contacto, 'es_llamada': False})
            r['con_contacto'] = r['con_contacto'] or con_contacto
            r['es_llamada'] = r['es_llamada'] or es_llamada

        medios_creados = 0
        for nombre, info in medios_info.items():
            medio, created = Medio.objects.get_or_create(
                cartera=cartera, nombre=nombre,
                defaults={'canal': info['canal'], 'es_llamada': info['es_llamada'], 'es_inbound': info['es_inbound']},
            )
            medio.canal = info['canal']
            medio.es_llamada = info['es_llamada']
            medio.es_inbound = info['es_inbound']
            medio.permite_manual = medio.calcular_permite_manual()
            medio.save(update_fields=['canal', 'es_llamada', 'es_inbound', 'permite_manual'])
            if created:
                medios_creados += 1

        resultados_creados, resultados_actualizados = 0, 0
        for (nombre, tipo_contacto), info in resultados_info.items():
            con_contacto = info['con_contacto']
            crea_compromiso = 'COMPROMISO DE PAGO' in nombre.upper()
            requiere_fecha = crea_compromiso or 'PAGO' in nombre.upper()
            # Nuevo Capital no reporta pagos por gestion; "paga tercero (o aval)" es la unica
            # excepcion acordada para llegar a Pagando sin pasar por el formulario de Pagos.
            efecto_pago = Resultado.EFECTO_PAGANDO if 'PAGA TERCERO' in nombre.upper() else ''

            resultado, created = Resultado.objects.get_or_create(
                cartera=cartera, nombre=nombre, tipo_contacto=tipo_contacto,
            )
            resultado.contactabilidad = Resultado.CON_CONTACTO if con_contacto else Resultado.SIN_CONTACTO
            resultado.crea_compromiso = crea_compromiso
            resultado.requiere_fecha_pago = requiere_fecha
            resultado.efecto_pago = efecto_pago
            if created:
                resultado.descarga_grabacion = info['es_llamada'] and con_contacto
                resultados_creados += 1
            else:
                resultados_actualizados += 1
            resultado.save()

        self.stdout.write(self.style.SUCCESS(
            f"Cartera 'Nuevo Capital': {medios_creados} medio(s) nuevo(s), "
            f"{resultados_creados} resultado(s) nuevo(s), {resultados_actualizados} actualizado(s)."
        ))
