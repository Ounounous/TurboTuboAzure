"""
Reporte "Gestiones Tanner a Omega": convierte las gestiones de un dia de la cartera Tanner al
formato que exige la carga masiva del sistema anterior (Omega) -- 8 columnas: RUT, FECHA, ACCION,
ESTADO, EMAIL, TELEFONO, COMENTARIO, SUBESTADO.

Se usa durante el traspaso de cobradores a produccion real: mientras Omega siga siendo el sistema
de referencia para algo, las gestiones que se hagan en TurboTubo tienen que poder subirse ahi
tambien, con el mismo vocabulario (Accion/Estado/SubEstado) que Omega ya usaba.

El mapeo CODIGO_A_OMEGA sale de cruzar dos fuentes reales, no de adivinar:
  1. La paleta oficial de Tanner (Paleta Respuesta Tanner-.xlsx), que trae (Accion, SubEstado,
     Estado) para cada resultado -- 613 filas, sin ambiguedad una vez fijado el Accion.
  2. Los 55.266 registros REALES que Omega genero en jun-jul 2026 (via export directo del
     sistema, no la paleta escrita): sirvieron para confirmar cual Accion "base" (LLAMADA,
     CORREO, WHATSAPP, SMS, IVR, TERRENO) usar cuando la paleta ofrece varias variantes
     (COMERCIAL/INFORMATIVO/RECIBIDO), y para verificar que los codigos SIN mapeo (ver
     CODIGOS_BLOQUEADOS) en efecto nunca ocurrieron -- 0 falsos bloqueos en 2 meses reales.

Los codigos que no tienen equivalente conocido en Omega (PAC/venta directa, resultados sin
uso real observado) se BLOQUEAN explicitamente: la fila se excluye del archivo y se avisa,
en vez de adivinar un Accion/Estado que Omega podria rechazar o cargar mal.
"""
import datetime

REPORT_TZ = datetime.timezone(datetime.timedelta(hours=-4))

# codigo Tanner -> (ACCION, SUBESTADO, ESTADO) exactos que espera la carga masiva de Omega.
# Construido y validado contra 55.266 gestiones reales (jun-jul 2026) -- ver investigacion en
# la conversacion de deploy. No modificar a mano sin volver a cruzar contra la paleta oficial.
CODIGO_A_OMEGA = {
    '100': ('LLAMADA', 'DIRECTO', 'PROMESA DE PAGO'),
    '101': ('LLAMADA', 'DIRECTO', 'INTENCION DE PAGO'),
    '104': ('LLAMADA', 'DIRECTO', 'EN PROCESO DE DACION'),
    '105': ('LLAMADA', 'DIRECTO', 'EN PROCESO DE RENEGOCIACION'),
    '106': ('LLAMADA AUTORIZADA', 'DIRECTO', 'INCAUTADO/REMATADO'),
    '107': ('LLAMADA', 'DIRECTO', 'NO ES LA FECHA QUE PACTO'),
    '108': ('LLAMADA', 'DIRECTO', 'DICE HABER PAGADO'),
    '109': ('LLAMADA', 'DIRECTO', 'SIN MODALIDAD DE PAGO'),
    '110': ('LLAMADA', 'DIRECTO', 'SINIESTRO'),
    '111': ('LLAMADA', 'DIRECTO', 'DESCONOCE DEUDA'),
    '112': ('LLAMADA', 'DIRECTO', 'PAGA TERCERO O AVAL'),
    '113': ('LLAMADA', 'DIRECTO', 'CESANTE'),
    '114': ('LLAMADA', 'DIRECTO', 'ENFERMEDAD / LICENCIA MEDICA'),
    '115': ('LLAMADA', 'DIRECTO', 'PROBLEMAS FINANCIEROS'),
    '116': ('LLAMADA', 'DIRECTO', 'SOLICITA LLAMADO POSTERIOR'),
    '117': ('LLAMADA', 'DIRECTO', 'NO QUIERE PAGAR'),
    '118': ('CORREO', 'DIRECTO', 'NEGOCIACIÓN POR EMAIL'),
    '119': ('WHATSAPP', 'DIRECTO', 'NEGOCIACIÓN POR WHATSAPP'),
    '120': ('LLAMADA', 'DIRECTO', 'VACACIONES'),
    '121': ('LLAMADA', 'DIRECTO', 'TRAMITANDO SEGURO'),
    '123': ('LLAMADA', 'DIRECTO', 'RESPUESTA AGRESIVA'),
    '124': ('LLAMADA', 'DIRECTO', 'OTROS'),
    '125': ('LLAMADA', 'DIRECTO', 'INTENCION DE DACION'),
    '126': ('LLAMADA', 'DIRECTO', 'INTENCION DE RENEGOCIACION'),
    '127': ('TERRENO', 'DIRECTO', 'CITACIÓN ENTREGADA A TITULAR'),
    '128': ('LLAMADA', 'DIRECTO', 'AUTORIZA LLAMADO POSTERIOR'),
    '137': ('LLAMADA', 'DIRECTO', 'INTERESADO EN VENTA DIRECTA'),
    '138': ('LLAMADA', 'DIRECTO', 'NO INTERESADO EN VENTA DIRECTA (OTROS)'),
    '139': ('LLAMADA', 'DIRECTO', 'REGULARIZARA POR OTRO MEDIO'),
    '140': ('LLAMADA', 'DIRECTO', 'VEHICULO COMO HERRAMIENTA DE TRABAJO'),
    '143': ('LLAMADA', 'DIRECTO', 'SIN RESPUESTA CLIENTE'),
    '146': ('LLAMADA', 'DIRECTO', 'DISCONFORMIDAD POR PRECIO'),
    '147': ('LLAMADA', 'DIRECTO', 'SIN RESPUESTA OFERTA'),
    '149': ('LLAMADA', 'DIRECTO', 'VEHICULO EN MAL ESTADO'),
    '200': ('LLAMADA', 'DIRECTO AVAL', 'PROMESA DE PAGO'),
    '201': ('LLAMADA', 'DIRECTO AVAL', 'INTENCION DE PAGO'),
    '204': ('LLAMADA', 'DIRECTO AVAL', 'EN PROCESO DE DACION'),
    '205': ('LLAMADA', 'DIRECTO AVAL', 'EN PROCESO DE RENEGOCIACION'),
    '206': ('LLAMADA AUTORIZADA', 'DIRECTO AVAL', 'INCAUTADO/REMATADO'),
    '207': ('LLAMADA', 'DIRECTO AVAL', 'NO ES LA FECHA QUE PACTO'),
    '208': ('LLAMADA', 'DIRECTO AVAL', 'DICE HABER PAGADO'),
    '209': ('LLAMADA', 'DIRECTO AVAL', 'SIN MODALIDAD DE PAGO'),
    '210': ('LLAMADA', 'DIRECTO AVAL', 'SINIESTRO'),
    '211': ('LLAMADA', 'DIRECTO AVAL', 'DESCONOCE DEUDA'),
    '213': ('LLAMADA', 'DIRECTO AVAL', 'CESANTE'),
    '214': ('CORREO', 'DIRECTO AVAL', 'ENFERMO'),
    '215': ('CORREO', 'DIRECTO AVAL', 'SIN DINERO'),
    '216': ('LLAMADA', 'DIRECTO AVAL', 'SOLICITA LLAMADO POSTERIOR'),
    '217': ('LLAMADA', 'DIRECTO AVAL', 'NO QUIERE PAGAR'),
    '218': ('LLAMADA', 'DIRECTO AVAL', 'VACACIONES'),
    '219': ('LLAMADA', 'DIRECTO AVAL', 'TRAMITANDO SEGURO'),
    '221': ('LLAMADA', 'DIRECTO AVAL', 'RESPUESTA AGRESIVA'),
    '222': ('LLAMADA', 'DIRECTO AVAL', 'OTROS'),
    '223': ('LLAMADA', 'DIRECTO AVAL', 'AUTORIZA LLAMADO POSTERIOR'),
    '300': ('LLAMADA', 'INDIRECTO', 'FALLECIDO'),
    '301': ('LLAMADA', 'INDIRECTO', 'PAGA TERCERO'),
    '302': ('LLAMADA', 'INDIRECTO', 'SE DEJA RECADO CON FAMILIAR'),
    '303': ('LLAMADA', 'INDIRECTO', 'SE DEJA RECADO CON TERCERO'),
    '304': ('LLAMADA RECIBIDA', 'INDIRECTO', 'TERCERO / FAMILIAR LLAMAR MAS TARDE'),
    '305': ('LLAMADA', 'INDIRECTO', 'TERCERO NO CONOCE A DEUDOR'),
    '306': ('TERRENO', 'INDIRECTO', 'DIRECCIÓN CONFIRMADA CON VECINO'),
    '307': ('TERRENO', 'INDIRECTO', 'CITACIÓN RECIBIDA POR TERCERO'),
    '308': ('TERRENO', 'INDIRECTO', 'TERCERO CONFIRMA CAMBIO DE DOMICILIO'),
    '309': ('LLAMADA', 'INDIRECTO', 'REHUSA ATENCIÓN'),
    '310': ('LLAMADA', 'INDIRECTO', 'AUTORIZA LLAMADO POSTERIOR'),
    '400': ('LLAMADA', 'SIN CONTACTO OPERADOR', 'SIN CONTACTO BUZON DE VOZ'),
    '401': ('LLAMADA', 'SIN CONTACTO OPERADOR', 'SIN CONTACTO NO CONTESTA'),
    '402': ('LLAMADA', 'SIN CONTACTO OPERADOR', 'TELEFONO NO CORRESPONDE'),
    '403': ('LLAMADA', 'SIN CONTACTO OPERADOR', 'CORTA LLAMADO'),
    '404': ('LLAMADA', 'SIN CONTACTO OPERADOR', 'BUSQUEDA DE DATOS'),
    '405': ('BUSQUEDA', 'SIN CONTACTO OPERADOR', 'SIN DATOS DEMOGRAFICOS'),
    '510': ('TERRENO', 'SIN CONTACTO TERRENO', 'DIRECCION NO CORRESPONDE'),
    '511': ('TERRENO', 'SIN CONTACTO TERRENO', 'DIRECCION INXISTENTE'),  # sic: typo real de la paleta oficial
    '512': ('TERRENO', 'SIN CONTACTO TERRENO', 'DIRECCIÓN DE CONOCIDO'),
    '513': ('TERRENO', 'SIN CONTACTO TERRENO', 'SIN MORADORES'),
    '514': ('TERRENO', 'SIN CONTACTO TERRENO', 'LUGAR INACCESIBLE'),
    '600': ('SMS', 'ACCION MASIVA', 'ENVIO SMS ENTREGADO'),
    '601': ('TERRENO', 'ACCION MASIVA', 'ENVIO CARTA ENTREGADO'),
    '602': ('CORREO', 'ACCION MASIVA', 'ENVIO EMAIL ENTREGADO'),
    '603': ('WHATSAPP', 'ACCION MASIVA', 'ENVIO WHATSAPP ENTREGADO'),
    '604': ('IVR', 'ACCION MASIVA', 'ENVIO IVR ENTREGADO'),
    '605': ('SMS', 'ACCION MASIVA', 'ENVIO SMS NO ENTREGADO'),
    '607': ('CORREO', 'ACCION MASIVA', 'ENVIO EMAIL NO ENTREGADO'),
    '608': ('WHATSAPP', 'ACCION MASIVA', 'ENVIO WHATSAPP NO ENTREGADO'),
    '609': ('IVR', 'ACCION MASIVA', 'ENVIO IVR INCOMPLETO'),
    '610': ('IVR', 'ACCION MASIVA', 'ENVIO IVR NO ENTREGADO'),
}

# Codigos del arbol Tanner de TurboTubo SIN equivalente confirmado en Omega: o son de
# campaña/backoffice que nunca pasan por aca (PAC/venta directa, permite_manual=False), o no
# aparecieron ni una vez en los 55.266 registros reales de 2 meses y no hay como confirmar el
# Accion/Estado correcto sin arriesgarse a que Omega los rechace o los cargue mal.
CODIGOS_BLOQUEADOS = {
    '102': 'YA PAGO', '103': 'PAGADO', '122': 'CONTINGENCIA',
    '202': 'YA PAGO (aval)', '203': 'PAGADO (aval)', '220': 'CONTINGENCIA (aval)',
    '500': 'SIN OPERADOR DISPONIBLE (DROP)', '501': 'SIN CONTACTO MAQUINA',
    '606': 'ENVIO CARTA NO ENTREGADO', '800': 'SIN GESTION',
    '129': 'PAC INTERESADO', '130': 'PAC CONTRATADO WEB', '131': 'PAC CONTRATADO FÍSICO',
    '132': 'PAGARA POR OTROS MEDIOS', '133': 'NO BANCARIZADO', '134': 'BANCO SIN CONVENIO',
    '135': 'NO INTERESADO SIN MOTIVO', '136': 'YA TIENE PAC',
    '141': 'INFORMACION ENVIADA A TANNER', '142': 'EN LICITACIÓN CONCESIONARIOS',
    '144': 'SIN GESTION CONCESIONARIO', '145': 'EN PROCESO REVISION VEHICULO',
    '148': 'SIN CONTACTO CONCESIONARIO', '150': 'VEHICULO CON DEMASIADO KM',
    '151': 'VEHICULO CON ANOTACIONES EN TRAMITE', '152': 'VEHICULO NO COMERCIAL',
    '153': 'VENDIDO',
}


def formatear_telefono(raw_number):
    """Formato 56XXXXXXXXX, igual que el reporte que Omega envia a Tanner. PENDIENTE DE
    CONFIRMAR: el archivo de ejemplo de carga a Omega (20260806_tanner_mail) no trae ningun caso
    con telefono relleno (todos eran de correo), asi que no hay como verificarlo contra un caso
    real. Si Omega rechaza el formato al subir la primera prueba, ajustar aca."""
    from .tanner_report import formatear_telefono as _fmt
    return _fmt(raw_number)


def rut_sin_dv(lead):
    """La plantilla de Omega lleva el RUT sin digito verificador, como numero."""
    try:
        return int(lead.rut)
    except (TypeError, ValueError):
        return lead.rut


def gestiones_del_dia(fecha):
    """Gestiones de TODA la cartera Tanner de un dia (igual que el reporte oficial a Tanner: sin
    acotar por subcartera, en el mismo huso horario)."""
    from .models import Action

    start = datetime.datetime.combine(fecha, datetime.time.min, tzinfo=REPORT_TZ)
    end = datetime.datetime.combine(fecha, datetime.time.max, tzinfo=REPORT_TZ)

    return Action.objects.filter(
        subcartera__cartera__nombre__iexact='Tanner',
        created_at__gte=start,
        created_at__lte=end,
    ).select_related('lead', 'medio', 'resultado', 'phone').order_by('created_at')


def construir_filas(actions):
    """Convierte gestiones a filas (RUT, FECHA, ACCION, ESTADO, EMAIL, TELEFONO, COMENTARIO,
    SUBESTADO). Devuelve (filas, codigos_bloqueados_encontrados): las gestiones con un resultado
    bloqueado se excluyen y se listan aparte, para avisar en vez de subir algo incorrecto."""
    filas = []
    bloqueados = []
    for action in actions:
        codigo = action.resultado.codigo
        mapeo = CODIGO_A_OMEGA.get(codigo)
        if mapeo is None:
            bloqueados.append(action)
            continue
        accion, subestado, estado = mapeo
        local_dt = action.created_at.astimezone(REPORT_TZ)
        filas.append([
            rut_sin_dv(action.lead),
            local_dt.replace(tzinfo=None),
            accion,
            estado,
            action.email or None,
            formatear_telefono(action.phone.phone_number) if action.phone else None,
            action.comment or '',
            subestado,
        ])
    return filas, bloqueados


def construir_workbook(fecha):
    """
    Arma el Excel de un dia completo (mismas 8 columnas, misma segunda hoja de excluidas que ya
    generaba la vista sincrona). Compartido por la vista de descarga y la tarea del worker, para
    que las dos rutas no puedan divergir en formato -- ver actions/tasks.py:generar_reporte_omega.

    Devuelve (wb|None, total_filas, total_excluidas). wb es None si no hay ninguna fila que
    incluir (ni gestiones ese dia, ni gestiones bloqueadas que avisar).
    """
    from openpyxl import Workbook

    actions = gestiones_del_dia(fecha)
    if not actions.exists():
        return None, 0, 0

    filas, bloqueados = construir_filas(actions)
    if not filas and not bloqueados:
        return None, 0, 0

    wb = Workbook()
    ws = wb.active
    ws.title = 'Datos'
    ws.append(['RUT', 'FECHA', 'ACCION', 'ESTADO', 'EMAIL', 'TELEFONO', 'COMENTARIO', 'SUBESTADO'])
    for fila in filas:
        ws.append(fila)
    for row in ws.iter_rows(min_row=2, min_col=2, max_col=2):
        row[0].number_format = 'm/d/yy h:mm'

    if bloqueados:
        ws2 = wb.create_sheet('Excluidas (sin mapeo a Omega)')
        ws2.append(['OP', 'RUT', 'Código resultado', 'Resultado', 'Motivo'])
        for a in bloqueados:
            ws2.append([
                a.lead.op, rut_sin_dv(a.lead), a.resultado.codigo, a.resultado.nombre,
                'Sin equivalente confirmado en Omega — súbela a mano si corresponde.',
            ])
        ws2.column_dimensions['D'].width = 40
        ws2.column_dimensions['E'].width = 55

    return wb, len(filas), len(bloqueados)
