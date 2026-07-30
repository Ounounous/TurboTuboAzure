"""
Construccion del ZIP de grabaciones. Vive aparte de las vistas para que lo pueda usar tanto la
vista (validacion rapida) como la tarea de Celery (generacion en el worker, sin bloquear la web).
"""
import datetime
import shutil
import tempfile
import zipfile

from django.db.models import Q

MAX_GRABACIONES = 500


def construir_zip_grabaciones(excel_fileobj, user, max_grabaciones=MAX_GRABACIONES):
    """
    Lee el Excel (cartera, subcartera, op, [fecha] por fila), junta las grabaciones de los leads
    VISIBLES para `user` y las escribe en un ZIP en disco (streaming, sin cargar cada audio entero
    a memoria). Devuelve (spool_tempfile_posicionado_al_inicio, total, errores).

    Si total == 0, igual devuelve el spool (con errores.txt adentro si hubo errores) para que el
    llamador decida; no lanza excepcion por "no encontro nada".
    """
    from openpyxl import load_workbook
    from django.utils.dateparse import parse_date
    from actions.models import CallRecording
    from demographics.views import find_lead
    from lead.permissions import scope_por_lead

    wb = load_workbook(excel_fileobj)
    sheet = wb.active

    spool = tempfile.SpooledTemporaryFile(max_size=8 * 1024 * 1024)
    errors = []
    total = 0
    excedido = False

    with zipfile.ZipFile(spool, 'w', zipfile.ZIP_DEFLATED) as zf:
        used_names = set()
        for row_number, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
            if excedido:
                break
            cartera, subcartera, op, fecha = (list(row) + [None] * 4)[:4]
            if not cartera or not subcartera or not op:
                errors.append(f"Fila {row_number}: faltan cartera, subcartera u OP.")
                continue

            lead = find_lead(cartera, subcartera, op)
            if not lead:
                errors.append(f"Fila {row_number}: no se encontró lead OP={op} en Cartera={cartera}, Subcartera={subcartera}.")
                continue

            # Alcance: solo grabaciones de leads visibles para el usuario que pidio el export.
            recordings = scope_por_lead(CallRecording.objects.filter(lead=lead), user)

            if fecha:
                parsed_date = None
                if isinstance(fecha, (datetime.date, datetime.datetime)):
                    parsed_date = fecha.date() if isinstance(fecha, datetime.datetime) else fecha
                else:
                    parsed_date = parse_date(str(fecha).strip())
                if parsed_date:
                    recordings = recordings.filter(
                        Q(call_date__date=parsed_date) | Q(call_date__isnull=True, created_at__date=parsed_date)
                    )
                else:
                    errors.append(f"Fila {row_number}: no se pudo interpretar la fecha '{fecha}', se ignoró el filtro.")

            for recording in recordings.iterator():
                if not recording.audio_file:
                    continue
                if total >= max_grabaciones:
                    excedido = True
                    break
                name = recording.audio_file.name.rsplit('/', 1)[-1]
                while name in used_names:
                    name = f"dup_{name}"
                used_names.add(name)
                with recording.audio_file.open('rb') as src, zf.open(name, 'w') as dst:
                    shutil.copyfileobj(src, dst, length=64 * 1024)
                total += 1

        if excedido:
            errors.append(
                f"Se alcanzó el máximo de {max_grabaciones} grabaciones por descarga. "
                "Acota el Excel (menos OP o un rango de fechas) y descarga por tramos."
            )
        if errors:
            zf.writestr('errores.txt', '\n'.join(errors))

    spool.seek(0)
    return spool, total, errors
